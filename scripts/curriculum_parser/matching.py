# -*- coding: utf-8 -*-
"""과목마스터 매칭 (v1 이관, graph_rag.norm 재사용).

과목마스터 시트: 정규화키 -> 공식과목명. 추천 자료에 없는 정규 편성 과목은
SUPPLEMENT_SUBJECTS 로 보충한다.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

from .cells import norm, repair_text, nfc

# 2022 개정 공식 과목 중 과목마스터(추천 데이터 유래)에 없는 과목 보충 목록.
SUPPLEMENT_SUBJECTS = [
    "공통국어1", "공통국어2", "공통수학1", "공통수학2", "공통영어1", "공통영어2",
    "기본수학1", "기본수학2", "기본영어1", "기본영어2",
    "통합사회1", "통합사회2", "통합과학1", "통합과학2", "과학탐구실험1", "과학탐구실험2",
    "정보", "한문", "진로와 직업", "인간과 경제활동",
    "스포츠 생활1", "스포츠 생활2",
    "일본 문화", "중국 문화", "독일어권 문화", "프랑스어권 문화",
    "스페인어권 문화", "러시아 문화", "아랍 문화", "베트남 문화",
]


def load_subject_master(master_path: Path) -> dict[str, str]:
    """과목마스터 시트에서 정규화키 -> 공식과목명 매핑을 만든다."""
    wb = openpyxl.load_workbook(master_path, read_only=True, data_only=True)
    if "과목마스터" not in wb.sheetnames:
        wb.close()
        raise SystemExit(f"과목마스터 시트가 없습니다: {master_path}")
    mapping: dict[str, str] = {}
    for row in wb["과목마스터"].iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        official = nfc(repair_text(row[0]))
        key = nfc(repair_text(row[1])) if len(row) > 1 and row[1] else norm(official)
        mapping[key] = official
        mapping[norm(official)] = official
    wb.close()
    for name in SUPPLEMENT_SUBJECTS:
        mapping.setdefault(norm(name), name)
    return mapping


def match_subject(name: str, master: dict[str, str]) -> str:
    """정제된 과목명 -> 공식과목명(미매칭 시 "")."""
    return master.get(norm(name), "")
