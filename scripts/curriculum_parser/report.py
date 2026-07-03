# -*- coding: utf-8 -*-
"""검증요약 / 정제로그 / 택N그룹 / 미매칭과목 시트 생성."""
from __future__ import annotations

from openpyxl import Workbook

from .model import HEADERS, SEMESTER_KEYS


def build_workbook(all_rows: list, results: list) -> Workbook:
    out = Workbook()
    ws = out.active
    ws.title = "고1편제표"
    ws.append(HEADERS)
    for r in all_rows:
        ws.append(r.to_output_row())

    _sheet_summary(out, results)
    _sheet_cleanlog(out, results)
    _sheet_groups(out, results)
    _sheet_unmatched(out, all_rows)
    _sheet_metanotes(out, results)
    return out


def _sheet_summary(out, results):
    ws = out.create_sheet("검증요약")
    ws.append(["학교", "파일", "시트", "시트확정", "과목행", "마스터매칭", "매칭률%",
               "학기커버리지", "택N그룹", "이슈"])
    for e in results:
        rate = round(e["matched"] / e["rows"] * 100, 1) if e["rows"] else 0
        flags = list(e["issues"])
        if e["rows"] and not (30 <= e["rows"] <= 250):
            flags.append(f"행수이상({e['rows']})")
        if e["rows"] and e["sem_cov"] == 0:
            flags.append("학기커버리지0(병합복원실패?)")
        if e.get("sheet_confirmed", "N") != "Y":
            flags.append("시트미확정")
        ws.append([e["school"], e["file"], e["sheet"], e.get("sheet_confirmed", "N"),
                   e["rows"], e["matched"], rate, e["sem_cov"], e["groups"],
                   "; ".join(flags)])


def _sheet_cleanlog(out, results):
    ws = out.create_sheet("정제로그")
    ws.append(["학교", "좌표/범위", "원본값", "정제값", "플래그"])
    for e in results:
        for log in e.get("logs", []):
            ws.append(log.as_row())


def _sheet_groups(out, results):
    ws = out.create_sheet("택N그룹")
    ws.append(["학교", "선택군ID", "택N", "학기", "멤버과목"])
    for e in results:
        for grp in e.get("group_objs", []):
            names = ", ".join(dict.fromkeys(m.name_raw for m in grp.members if m.name_raw))
            ws.append([e["school"], grp.choice_id, grp.take_n, grp.semester, names])


def _sheet_unmatched(out, all_rows):
    ws = out.create_sheet("미매칭과목")
    ws.append(["학교", "구분", "교과군", "과목유형", "과목명"])
    for r in all_rows:
        if not r.official_name:
            ws.append([r.school, r.section, r.subject_group, r.subject_type, r.raw_name])


def _sheet_metanotes(out, results):
    ws = out.create_sheet("하단안내문")
    ws.append(["학교", "원문(본표제외·보존)"])
    for e in results:
        for note in e.get("meta_notes", []):
            ws.append([e["school"], note])
