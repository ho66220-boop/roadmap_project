# -*- coding: utf-8 -*-
"""병합셀 복원 → 밀집 그리드 (v2 핵심 신규).

v1 결함: read_only 로 읽어 세로 병합의 앵커 외 셀이 None → 학기/학점 유실.
v2: ws.merged_cells.ranges 를 순회해 앵커값을 전파(propagate)하거나,
줄바꿈 적층 학점("4\\n4\\n4")을 행 오프셋별로 분배(distribute)한다.

반환 grid 는 0-based list[list]. resolve_merge_semantics 는 worksheet 객체를
받는 순수 함수라 테스트에서 임시 파일이 불필요하다(make_ws 헬퍼로 인메모리 생성).
"""
from __future__ import annotations

from openpyxl.utils import range_boundaries

from .cells import numeric_tokens
from .model import CleanLog


def _coord(col: int, row: int) -> str:
    from openpyxl.utils import get_column_letter
    return f"{get_column_letter(col)}{row}"


def resolve_merge_semantics(ws, school: str = "") -> tuple[list, set, list]:
    """병합을 복원한 밀집 그리드를 만든다.

    반환:
      grid: list[list]  0-based 값 그리드
      nonanchor: set[(r0, c0)]  병합의 비앵커 셀 좌표(과목명 중복 방지에 사용)
      logs: list[CleanLog]  분배 불일치(stack_mismatch) 기록
    """
    maxr = ws.max_row or 0
    maxc = ws.max_column or 0
    grid = [[ws.cell(row=r, column=c).value for c in range(1, maxc + 1)]
            for r in range(1, maxr + 1)]
    nonanchor: set = set()
    logs: list = []

    for merge in ws.merged_cells.ranges:
        c1, r1, c2, r2 = range_boundaries(str(merge))  # 1-based, 포함
        anchor = grid[r1 - 1][c1 - 1]

        # 비앵커 좌표 기록
        for rr in range(r1, r2 + 1):
            for cc in range(c1, c2 + 1):
                if (rr, cc) != (r1, c1):
                    nonanchor.add((rr - 1, cc - 1))

        if anchor is None:
            continue

        toks, nums = numeric_tokens(anchor)
        span = r2 - r1 + 1

        # 분배: 단일 열 세로 병합 + 숫자 토큰 2개 이상 & 행 span 이하
        if c1 == c2 and span >= 2 and 2 <= len(nums) <= span:
            vals = nums + [nums[-1]] * (span - len(nums))
            for off in range(span):
                grid[r1 - 1 + off][c1 - 1] = vals[off]
            if len(nums) != span:
                logs.append(CleanLog(
                    school=school, coord=str(merge), raw=str(anchor),
                    cleaned=" / ".join(vals),
                    flag="stack_mismatch",
                ))
        else:
            # 전파: 앵커값을 병합 전 셀에 복사
            for rr in range(r1, r2 + 1):
                for cc in range(c1, c2 + 1):
                    grid[rr - 1][cc - 1] = anchor

    return grid, nonanchor, logs
