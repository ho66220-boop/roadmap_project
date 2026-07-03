# -*- coding: utf-8 -*-
"""편제표 결정론적 파서 v2 패키지.

공개 API:
  parse_workbook(ws, school, source_file, master) -> SheetResult
      worksheet 객체를 받는 순수 함수(임시 파일 불필요, 테스트 친화적).
  parse_all(source_dir, master_path, sheetmap_records) -> (rows, results)
"""
from __future__ import annotations

import re
from pathlib import Path

from . import cells, header, loader, matching
from .choice import RawRow, assign_choice_groups
from .classify import META_PATTERN, classify_row
from .model import (COHORT_YEAR, HEADERS, SEMESTER_KEYS, CleanLog, ParsedRow,
                    SheetResult)

__all__ = ["parse_workbook", "parse_all", "HEADERS", "SEMESTER_KEYS", "COHORT_YEAR"]


def _get(row: list, idx):
    if idx is None or idx < 0 or idx >= len(row):
        return None
    return row[idx]


def _build_raw_rows(grid, col, nonanchor, school, logs) -> tuple[list, list]:
    """그리드 데이터 영역 → RawRow 목록 + 학기 키(존재하는 것만, 정규 순서)."""
    sem_cols = col["semesters"]                       # {semkey: colidx}
    sem_keys = [k for k in SEMESTER_KEYS if k in sem_cols]
    data_start = col["data_start"]
    name_col = col["name"]
    sec_col, grp_col, type_col = col["section"], col["group"], col["type"]
    base_col, run_col = col["base"], col["run"]

    raw_rows: list = []
    running_section = ""
    running_group = ""

    for gi in range(data_start, len(grid)):
        row = grid[gi]
        raw = RawRow(idx=gi)

        sec_val = _get(row, sec_col)
        if sec_val and str(sec_val).strip():
            running_section = cells.normalize_section(str(sec_val))
        raw.section = running_section

        grp_val = _get(row, grp_col)
        if grp_val and str(grp_val).strip():
            running_group = cells.normalize_group(str(grp_val))
        raw.group = running_group

        type_val = _get(row, type_col)
        raw.subtype = cells.nfc(cells.repair_text(type_val)) if type_val else ""

        name_val = _get(row, name_col)
        raw.name_raw = cells.nfc(cells.repair_text(name_val)) if name_val else ""
        raw.nonanchor_name = (gi, name_col) in nonanchor

        raw.base = cells.to_credit(_get(row, base_col))
        raw.run = cells.to_credit(_get(row, run_col))
        raw.is_skip = bool(raw.name_raw) and cells.is_skip_name(raw.name_raw)

        # 학기열: 숫자 학점 / 택N 마커 / 교차집중 / 비앵커 여부
        for k in sem_keys:
            ci = sem_cols[k]
            cellval = _get(row, ci)
            if cellval is None:
                continue
            text = str(cellval)
            val, flag = cells.clean_value(cellval)
            if isinstance(val, (int, float)) and val > 0:
                raw.sem_num[k] = float(val)
                if flag and flag not in ("nfc",):
                    logs.append(CleanLog(school, cells_coord(ci, gi), text,
                                         str(val), flag))
            m = cells.TAKE_N.search(text)
            if m:
                raw.sem_marker[k] = int(m.group(1))
                if (gi, ci) in nonanchor:
                    raw.sem_nonanchor.add(k)
            if cells.CROSS_CONCENTRATE.search(text):
                raw.cross_grade = k.split("-")[0]

        # 공동교육과정 등 학기열 서술에서 명시 학기 복원(직접 학점/마커 없을 때만)
        if not raw.sem_num and not raw.sem_marker and not raw.cross_grade:
            for k in sem_keys:
                cellval = _get(row, sem_cols[k])
                joints = cells.parse_joint_semesters(cellval) if cellval else []
                if joints:
                    fb = raw.run or raw.base or 0.0
                    for jk in joints:
                        if jk in sem_cols:
                            raw.sem_num[jk] = fb
                    break

        # 과목명 접미 (택N)
        mm = re.search(r"\(\s*택\s*(\d+)\s*\)\s*$", raw.name_raw)
        if mm:
            raw.name_suffix_take = int(mm.group(1))

        # 기타열(학기/과목명 아닌 열)의 택N
        if not raw.name_suffix_take and not raw.sem_marker:
            for ci in range(len(row)):
                if ci == name_col or ci in sem_cols.values():
                    continue
                cv = _get(row, ci)
                if cv and cells.TAKE_N.search(str(cv)):
                    raw.other_take = int(cells.TAKE_N.search(str(cv)).group(1))
                    break

        raw_rows.append(raw)

    return raw_rows, sem_keys


def cells_coord(col_idx0: int, row_idx0: int) -> str:
    from openpyxl.utils import get_column_letter
    return f"{get_column_letter(col_idx0 + 1)}{row_idx0 + 1}"


def _collect_meta_notes(grid, data_start) -> list[str]:
    notes: list[str] = []
    seen = set()
    for gi in range(data_start, len(grid)):
        for cv in grid[gi]:
            if not cv:
                continue
            text = cells.nfc(str(cv)).strip()
            if len(text) >= 25 and META_PATTERN.search(text):
                key = text[:60]
                if key not in seen:
                    seen.add(key)
                    notes.append(text)
    return notes


def parse_workbook(ws, school: str, source_file: str, master: dict) -> SheetResult:
    """worksheet 하나를 파싱한다(순수 함수)."""
    result = SheetResult()
    grid, nonanchor, merge_logs = loader.resolve_merge_semantics(ws, school)
    result.logs.extend(merge_logs)

    header_idx = header.find_header(grid)
    if header_idx is None:
        result.issues.append("헤더 행을 찾지 못함")
        return result
    col = header.build_column_map(grid, header_idx)
    if col is None:
        result.issues.append("열 구조 해석 실패(과목명/교과군/학년 열 누락)")
        return result

    raw_rows, sem_keys = _build_raw_rows(grid, col, nonanchor, school, result.logs)

    groups, _ = assign_choice_groups(raw_rows, sem_keys, school)
    result.groups.extend(groups)

    result.meta_notes.extend(_collect_meta_notes(grid, col["data_start"]))

    sheet_title = ws.title
    for raw in raw_rows:
        if not raw.name_raw:
            continue
        all_zero = not any(v > 0 for v in raw.credits.values()) and not raw.base and not raw.run
        label = classify_row(raw.name_raw, raw.base, raw.run, all_zero)
        if label in ("empty", "skip", "legend"):
            continue
        if label == "meta":
            if raw.name_raw not in result.meta_notes:
                result.meta_notes.append(raw.name_raw)
            continue
        if raw.nonanchor_name:
            continue

        if label == "cross":
            sides = [s for s in re.split(r"↔", raw.name_raw) if s.strip()]
            for side in sides:
                _emit(result, raw, side, school, source_file, sheet_title, master, cross=True)
            continue

        for piece in cells.split_subject_cell(raw.name_raw):
            if cells.is_legend(piece):
                continue
            _emit(result, raw, piece, school, source_file, sheet_title, master)

    if not result.rows:
        result.issues.append("파싱된 과목 행이 없음")
    return result


def _emit(result, raw, piece, school, source_file, sheet_title, master, cross=False):
    cleaned = cells.strip_markers(piece)
    if not cleaned:
        return
    official = matching.match_subject(cleaned, master)
    result.rows.append(ParsedRow(
        school=school,
        section=raw.section,
        subject_group=raw.group,
        subject_type=raw.subtype,
        raw_name=piece.strip(),
        official_name=official,
        base_credit=raw.base,
        run_credit=raw.run,
        credits=dict(raw.credits),
        source_file=source_file,
        source_sheet=sheet_title,
        choice_id=raw.choice_id,
        take_n=raw.take_n,
        cross=cross,
    ))


def parse_all(source_dir: Path, master: dict, sheetmap_records: list) -> tuple[list, list]:
    """시트맵 레코드대로 전 학교를 파싱한다.

    반환: (all_rows, results) — results 는 학교별 dict(진단·집계 포함).
    """
    import openpyxl
    source_dir = Path(source_dir)
    all_rows: list = []
    results: list = []

    for rec in sheetmap_records:
        school = rec["school"]
        entry = {"school": school, "file": "", "sheet": "", "rows": 0, "matched": 0,
                 "sem_cov": 0, "groups": 0, "sheet_confirmed": rec.get("sheet_confirmed", "N"),
                 "issues": [], "logs": [], "meta_notes": [], "group_objs": []}
        results.append(entry)

        path = _pick_file(source_dir, rec["file_keyword"])
        if path is None:
            entry["issues"].append(f"파일 없음(키워드: {rec['file_keyword']})")
            continue
        entry["file"] = path.name

        wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
        sheet = _pick_sheet(wb, rec.get("sheet_name", ""))
        if sheet is None:
            entry["issues"].append(f"시트 선택 실패(후보: {wb.sheetnames})")
            wb.close()
            continue
        entry["sheet"] = sheet

        res = parse_workbook(wb[sheet], school, path.name, master)
        wb.close()

        entry["rows"] = len(res.rows)
        entry["matched"] = sum(1 for r in res.rows if r.official_name)
        entry["sem_cov"] = sum(1 for r in res.rows for k in SEMESTER_KEYS
                               if r.credits.get(k, 0) and r.credits[k] > 0)
        entry["groups"] = len(res.groups)
        entry["issues"].extend(res.issues)
        entry["logs"] = res.logs
        entry["meta_notes"] = res.meta_notes
        entry["group_objs"] = res.groups
        all_rows.extend(res.rows)

    return all_rows, results


def _pick_file(source_dir: Path, keyword: str):
    candidates = [f for f in source_dir.glob("*.xlsx") if keyword in f.name]
    if not candidates:
        return None
    for f in candidates:
        if "신입생" in f.name or "입학생" in f.name:
            return f
    return candidates[0]


def _pick_sheet(wb, sheet_name: str):
    if sheet_name:
        if sheet_name in wb.sheetnames:
            return sheet_name
        for name in wb.sheetnames:
            if name.strip() == sheet_name.strip():
                return name
        return None
    candidates = [n for n in wb.sheetnames if "유의" not in n]
    return candidates[0] if len(candidates) == 1 else None
