"""Create a minimal public sample roadmap workbook.

The private source workbook is intentionally not committed. This script creates
a tiny synthetic workbook with the sheet names and columns expected by the MVP.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook


OUTPUT = Path(__file__).resolve().parents[1] / "data" / "sample_roadmap.xlsx"


def main() -> int:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()

    school = workbook.active
    school.title = "학교편제표"
    school.append(["학교명", "학년도", "교과군", "과목유형", "과목명", "공식코드", "학점", "1-1", "1-2", "2-1", "2-2", "3-1", "3-2"])
    school_rows = [
        ["울산고", 2026, "수학", "공통", "공통수학1", "MATH-COM-1", 4, 4, 0, 0, 0, 0, 0],
        ["울산고", 2026, "수학", "공통", "공통수학2", "MATH-COM-2", 4, 0, 4, 0, 0, 0, 0],
        ["울산고", 2026, "과학", "공통", "통합과학1", "SCI-COM-1", 4, 4, 0, 0, 0, 0, 0],
        ["울산고", 2026, "과학", "공통", "통합과학2", "SCI-COM-2", 4, 0, 4, 0, 0, 0, 0],
        ["울산고", 2026, "수학", "일반 선택", "대수", "MATH-ALG", 4, 0, 0, 4, 0, 0, 0],
        ["울산고", 2026, "수학", "일반 선택", "확률과 통계", "MATH-STA", 4, 0, 0, 0, 4, 0, 0],
        ["울산고", 2026, "수학", "진로 선택", "미적분Ⅰ", "MATH-CAL1", 4, 0, 0, 4, 0, 0, 0],
        ["울산고", 2026, "수학", "진로 선택", "미적분Ⅱ", "MATH-CAL2", 4, 0, 0, 0, 4, 0, 0],
        ["울산고", 2026, "과학", "일반 선택", "물리학", "SCI-PHY", 4, 0, 0, 4, 0, 0, 0],
        ["울산고", 2026, "과학", "일반 선택", "화학", "SCI-CHE", 4, 0, 0, 0, 4, 0, 0],
        ["울산고", 2026, "정보", "진로 선택", "인공지능 기초", "INFO-AI", 3, 0, 0, 0, 0, 3, 0],
        ["울산고", 2026, "정보", "진로 선택", "데이터 과학", "INFO-DATA", 3, 0, 0, 0, 0, 0, 3],
    ]
    for row in school_rows:
        school.append(row)

    major = workbook.create_sheet("로드맵DB_v2_대학트랙집계")
    major.append(["계열", "학과", "과목코드", "과목명", "교과군", "과목유형", "추천구분", "우선순위", "대학수", "비고1", "비고2", "근거"])
    major_rows = [
        ["공학", "컴퓨터공학", "MATH-ALG", "대수", "수학", "일반 선택", "핵심", 10, 5, "", "", "컴퓨터공학 전공 기초 수학 역량과 연결되는 샘플 근거"],
        ["공학", "컴퓨터공학", "MATH-STA", "확률과 통계", "수학", "일반 선택", "핵심", 20, 4, "", "", "데이터 분석과 통계적 사고 역량을 위한 샘플 근거"],
        ["공학", "컴퓨터공학", "MATH-CAL1", "미적분Ⅰ", "수학", "진로 선택", "권장", 30, 3, "", "", "알고리즘과 공학 수학 학습 기반을 위한 샘플 근거"],
        ["공학", "컴퓨터공학", "INFO-AI", "인공지능 기초", "정보", "진로 선택", "권장", 40, 2, "", "", "AI 및 소프트웨어 진로 탐색을 위한 샘플 근거"],
        ["공학", "컴퓨터공학", "INFO-DATA", "데이터 과학", "정보", "진로 선택", "권장", 50, 2, "", "", "데이터 기반 문제 해결 역량을 위한 샘플 근거"],
        ["자연", "생명공학", "SCI-CHE", "화학", "과학", "일반 선택", "핵심", 10, 5, "", "", "생명공학 실험과 물질 이해를 위한 샘플 근거"],
        ["자연", "생명공학", "SCI-COM-2", "통합과학2", "과학", "공통", "권장", 20, 3, "", "", "생명과학 기초 이해를 위한 샘플 근거"],
    ]
    for row in major_rows:
        major.append(row)

    alias = workbook.create_sheet("과목별칭")
    alias.append(["구분", "원문", "정규명"])
    alias_rows = [
        ["샘플", "미적분1", "미적분Ⅰ"],
        ["샘플", "미적분I", "미적분Ⅰ"],
        ["샘플", "인공지능", "인공지능 기초"],
    ]
    for row in alias_rows:
        alias.append(row)

    workbook.save(OUTPUT)
    print(f"Wrote {OUTPUT}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
