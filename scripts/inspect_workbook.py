"""Inspect the base roadmap workbook and print sheet dimensions.

Usage:
    python scripts/inspect_workbook.py ../울산고교_과목로드맵_수정.xlsx
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl


def inspect_workbook(path: Path) -> dict:
    workbook = openpyxl.load_workbook(path, data_only=True)
    return {
        "workbook": path.name,
        "sheets": [
            {
                "name": sheet.title,
                "rows": sheet.max_row,
                "columns": sheet.max_column,
                "headers": [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))],
            }
            for sheet in workbook.worksheets
        ],
    }


def main() -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    if len(sys.argv) != 2:
        print("Usage: python scripts/inspect_workbook.py <workbook.xlsx>", file=sys.stderr)
        return 2

    path = Path(sys.argv[1])
    if not path.exists():
        print(f"Workbook not found: {path}", file=sys.stderr)
        return 1

    print(json.dumps(inspect_workbook(path), ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
