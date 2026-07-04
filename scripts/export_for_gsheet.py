# -*- coding: utf-8 -*-
"""구글 시트 웹앱용 통합 업로드 엑셀을 생성한다.

전처리 편제표(curriculum_g1_2026.xlsx)와 로드맵 엑셀의 추천 데이터 시트를
하나의 파일로 합쳐, 구글 드라이브에 올려 시트로 변환하면 바로 쓸 수 있게 한다.

사용 예:
    python scripts/export_for_gsheet.py
"""
from __future__ import annotations

import argparse
import csv
from pathlib import Path
import sys

import openpyxl
from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_MASTER = ROOT.parent / "울산고교_과목로드맵_수정.xlsx"
DEFAULT_CURRICULUM = ROOT / "data" / "curriculum_g1_2026.xlsx"
DEFAULT_OUT = ROOT / "data" / "gsheet_upload.xlsx"
DEFAULT_TARGET_SCHOOLS = ROOT / "data" / "target_schools.csv"

# (원본 파일 구분, 시트명)
SHEETS = [
    ("curriculum", "고1편제표"),
    ("master", "학과추천"),
    ("master", "대학트랙"),
    ("master", "대학제시율_학과"),
    ("master", "대학제시율_계열"),
    ("master", "과목마스터"),
]

# 노출 필터를 적용할 시트(학교 단위 행 필터). 그 외 시트는 무필터.
FILTERED_SHEETS = {"고1편제표"}
SCHOOL_COLUMN = "학교"


def read_rows(src_ws) -> list[list]:
    """완전히 빈 행은 제외하고 값 행만 반환한다."""
    rows = []
    for row in src_ws.iter_rows(values_only=True):
        if any(c not in (None, "") for c in row):
            rows.append(["" if c is None else c for c in row])
    return rows


def load_target_schools(path: Path) -> set[str]:
    """target_schools.csv 의 school 컬럼(노출 대상 학교명) 집합을 반환."""
    if not path.exists():
        return set()
    with path.open(encoding="utf-8-sig", newline="") as f:
        return {(row.get("school") or "").strip()
                for row in csv.DictReader(f) if (row.get("school") or "").strip()}


def copy_sheet(src_ws, out_wb, sheet_name: str, allowed_schools: set[str] | None = None) -> int:
    """src_ws 를 out_wb 에 sheet_name 으로 복사. allowed_schools 가 주어지면
    헤더의 '학교' 컬럼 값이 그 집합에 속한 행만 남긴다(헤더는 항상 유지).
    반환값: 데이터 행 수(헤더 제외)."""
    rows = read_rows(src_ws)
    ws = out_wb.create_sheet(title=sheet_name)
    if not rows:
        return 0
    header, data_rows = rows[0], rows[1:]
    if allowed_schools is not None and SCHOOL_COLUMN in header:
        col = header.index(SCHOOL_COLUMN)
        data_rows = [r for r in data_rows if r[col] in allowed_schools]
    ws.append(header)
    for r in data_rows:
        ws.append(r)
    return len(data_rows)


def main() -> int:
    parser = argparse.ArgumentParser(description="구글 시트 업로드용 통합 엑셀 생성")
    parser.add_argument("--master", default=str(DEFAULT_MASTER))
    parser.add_argument("--curriculum", default=str(DEFAULT_CURRICULUM))
    parser.add_argument("--out", default=str(DEFAULT_OUT))
    parser.add_argument("--target-schools", default=str(DEFAULT_TARGET_SCHOOLS),
                        help="--target-only 사용 시 노출 대상 학교 CSV(school 컬럼)")
    parser.add_argument("--target-only", action="store_true",
                        help="고1편제표를 target_schools.csv 학교만으로 필터(기본은 파싱된 전 학교 노출)")
    args = parser.parse_args()

    sources = {}
    for key, path in (("master", Path(args.master)), ("curriculum", Path(args.curriculum))):
        if not path.exists():
            print(f"파일이 없습니다: {path}", file=sys.stderr)
            return 1
        sources[key] = openpyxl.load_workbook(path, read_only=True, data_only=True)

    allowed_schools = None
    if args.target_only:
        allowed_schools = load_target_schools(Path(args.target_schools))
        if not allowed_schools:
            print(f"노출 대상 학교 CSV가 비어있습니다: {args.target_schools}", file=sys.stderr)
            return 1
        print(f"노출 필터 적용: {len(allowed_schools)}개교 (기본은 전 학교 노출)")

    out = Workbook()
    out.remove(out.active)
    for source_key, sheet_name in SHEETS:
        wb = sources[source_key]
        if sheet_name not in wb.sheetnames:
            print(f"시트가 없습니다: {sheet_name} ({source_key})", file=sys.stderr)
            return 1
        filt = allowed_schools if sheet_name in FILTERED_SHEETS else None
        rows = copy_sheet(wb[sheet_name], out, sheet_name, filt)
        print(f"  {sheet_name}: {rows}행")

    out_path = Path(args.out)
    out.save(out_path)
    print(f"\n저장: {out_path}")
    print("구글 드라이브에 업로드 후 'Google Sheets로 열기'로 변환해서 사용하세요.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
