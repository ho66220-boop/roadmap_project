# -*- coding: utf-8 -*-
"""완전 가상 샘플 데이터셋 생성기(결정론적, openpyxl만 사용).

목적: 원본 데이터(저작권·개인정보 문제로 비공개) 없이도 클론만으로
build_curriculum.py -> server.py 파이프라인 전체를 체험할 수 있게 한다.
학교·대학·학과·과목명은 전부 가상이며(과목명은 2022 개정 공식 과목명 사용),
편제표 원자료에는 실제 파일에서 흔한 오염 요소(병합셀·택N·다과목 셀·각주·
온라인공동교육과정 접미·정제 대상 오염값)를 의도적으로 넣어 파서 동작을
직접 관찰할 수 있게 했다.

산출:
  data/sample/master.xlsx               학과추천/대학트랙/대학제시율_학과/
                                         대학제시율_계열/과목마스터 5개 시트
  data/sample/curriculum_source/*.xlsx  가상 학교 2곳(한빛고, 미리내고) 편제표 원자료
  data/sample/sheetmap.csv              위 두 학교 레코드

사용:
    python scripts/make_sample_data.py
    python scripts/build_curriculum.py --source-dir data/sample/curriculum_source \
        --master data/sample/master.xlsx --config data/sample/sheetmap.csv \
        --out data/sample/curriculum_g1_2026.xlsx
    python server.py --workbook data/sample/master.xlsx \
        --curriculum data/sample/curriculum_g1_2026.xlsx
"""
from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
SAMPLE_DIR = ROOT / "data" / "sample"
SOURCE_DIR = SAMPLE_DIR / "curriculum_source"

# --------------------------------------------------------------------------
# 과목마스터 — 2022 개정 공식 과목명 38개(공공 정보). 일부러 "경제 수학"은
# 빼 두어 편제표 매칭 시 미매칭 과목 1건이 생기게 한다(미매칭과목 시트 시연).
# --------------------------------------------------------------------------
SUBJECT_MASTER = [
    "문학", "독서와 작문", "화법과 언어", "주제 탐구 독서",
    "대수", "미적분Ⅰ", "미적분Ⅱ", "기하", "확률과 통계", "인공지능 수학",
    "영어Ⅰ", "영어Ⅱ", "영어 독해와 작문",
    "세계사", "동아시아 역사 기행", "한국지리 탐구", "사회와 문화", "정치와 법", "경제",
    "한국사1", "한국사2",
    "물리학", "화학", "생명과학", "지구과학", "과학과제 연구", "역학과 에너지",
    "정보", "인공지능 기초", "데이터 과학", "소프트웨어와 생활",
    "음악", "미술", "체육1",
    "중국어", "일본어",
    "심리학", "논리학",
]

# 학과추천 — 가상 학과 3개(계열/일반/진로/융합)
DEPT_RECS = {
    "컴퓨터공학과": {
        "track": "공학",
        "일반": ["확률과 통계", "영어 독해와 작문"],
        "진로": ["미적분Ⅰ", "미적분Ⅱ", "정보", "인공지능 기초", "데이터 과학"],
        "융합": ["소프트웨어와 생활", "인공지능 수학"],
    },
    "간호학과": {
        "track": "의약",
        "일반": ["화법과 언어", "독서와 작문"],
        "진로": ["생명과학", "화학", "심리학"],
        "융합": ["지구과학", "논리학"],
    },
    "경영학과": {
        "track": "사회",
        "일반": ["경제", "사회와 문화"],
        "진로": ["정치와 법", "세계사"],
        "융합": ["논리학", "한국지리 탐구"],
    },
}

# 대학트랙 — 가상 대학 3개(가온대/누리대/한별대). 누리대·한별대 컴퓨터공학
# 계열 행에는 일부러 "기하 또는 미적분Ⅱ"(복수 과목 표기), "수학"(교과군 포괄
# 표현), "미적분"(구 교육과정 명칭) 을 넣어 _expand_track_subject 의 별칭/
# 교과군 확장 로직을 시연한다.
# (대학, 모집단위, 과목, 구분, 우선순위, 비고)
UNI_TRACKS = [
    ("가온대", "컴퓨터공학과", "정보", "핵심", 1, "2025 시행계획 기준(샘플)"),
    ("가온대", "컴퓨터공학과", "미적분Ⅱ", "핵심", 2, "2025 시행계획 기준(샘플)"),
    ("가온대", "컴퓨터공학과", "확률과 통계", "권장", 3, "2025 시행계획 기준(샘플)"),
    ("누리대", "컴퓨터공학전공", "기하 또는 미적분Ⅱ", "핵심", 1, "복수 과목 중 택1 표기 샘플"),
    ("누리대", "컴퓨터공학전공", "데이터 과학", "권장", 2, "2025 시행계획 기준(샘플)"),
    ("한별대", "컴퓨터공학과", "수학", "핵심", 1, "교과군 포괄 표기 샘플"),
    ("한별대", "컴퓨터공학과", "미적분", "권장", 2, "구 교육과정 명칭 표기 샘플"),
    ("가온대", "간호학과", "생명과학", "핵심", 1, "2025 시행계획 기준(샘플)"),
    ("가온대", "간호학과", "화학", "권장", 2, "2025 시행계획 기준(샘플)"),
    ("누리대", "간호학전공", "화학", "핵심", 1, "2025 시행계획 기준(샘플)"),
    ("한별대", "간호학과", "심리학", "권장", 1, "2025 시행계획 기준(샘플)"),
    ("가온대", "경영학과", "경제", "핵심", 1, "2025 시행계획 기준(샘플)"),
    ("가온대", "경영학과", "사회와 문화", "권장", 2, "2025 시행계획 기준(샘플)"),
    ("누리대", "경영학부", "정치와 법", "핵심", 1, "2025 시행계획 기준(샘플)"),
    ("한별대", "경영학과", "논리학", "권장", 1, "2025 시행계획 기준(샘플)"),
]

# 대학제시율 — (대상, 과목, 대학수, 비고, 제시율%)
RATE_BY_DEPT = [
    ("컴퓨터공학과", "정보", 12, "", 85),
    ("컴퓨터공학과", "미적분Ⅱ", 10, "", 70),
    ("컴퓨터공학과", "확률과 통계", 8, "", 55),
    ("간호학과", "생명과학", 9, "", 80),
    ("간호학과", "화학", 7, "", 65),
    ("경영학과", "경제", 11, "", 75),
    ("경영학과", "정치와 법", 6, "", 50),
]
RATE_BY_TRACK = [
    ("공학", "미적분Ⅱ", 30, "", 72),
    ("공학", "정보", 28, "", 68),
    ("의약", "화학", 22, "", 77),
    ("의약", "생명과학", 25, "", 81),
    ("사회", "경제", 20, "", 60),
    ("사회", "사회와 문화", 18, "", 58),
]

SHEETMAP_RECORDS = [
    {"school": "한빛고", "file_keyword": "한빛고 2026학년도", "sheet_name": "",
     "sheet_confirmed": "Y", "note": "가상 샘플(전수 자동선정)"},
    {"school": "미리내고", "file_keyword": "미리내고 2026학년도", "sheet_name": "",
     "sheet_confirmed": "Y", "note": "가상 샘플(세부교과목 헤더변형)"},
]

# --------------------------------------------------------------------------
# master.xlsx 생성
# --------------------------------------------------------------------------

def build_master_workbook() -> Workbook:
    wb = Workbook()

    ws = wb.active
    ws.title = "학과추천"
    ws.append(["학과", "계열", "구분", "과목"])
    for dept, cfg in DEPT_RECS.items():
        track = cfg["track"]
        for kind in ("일반", "진로", "융합"):
            for subject in cfg[kind]:
                ws.append([dept, track, kind, subject])

    ws = wb.create_sheet("대학트랙")
    ws.append(["대학", "모집단위", "과목", "구분", "우선순위", "비고"])
    for row in UNI_TRACKS:
        ws.append(list(row))

    ws = wb.create_sheet("대학제시율_학과")
    ws.append(["대상학과", "과목", "대학수", "비고", "제시율"])
    for row in RATE_BY_DEPT:
        ws.append(list(row))

    ws = wb.create_sheet("대학제시율_계열")
    ws.append(["계열", "과목", "대학수", "비고", "제시율"])
    for row in RATE_BY_TRACK:
        ws.append(list(row))

    ws = wb.create_sheet("과목마스터")
    ws.append(["공식과목명"])
    for name in SUBJECT_MASTER:
        ws.append([name])

    return wb


# --------------------------------------------------------------------------
# 편제표 원자료 생성 — 실제 원자료 양식(1행 제목, 3행 헤더, 5행 학기라벨,
# 6행부터 데이터) 재현 + 병합셀/택N/다과목 셀/각주/정제 오염값 포함.
# --------------------------------------------------------------------------
SEM_COL = {"1-1": 6, "1-2": 7, "2-1": 8, "2-2": 9, "3-1": 10, "3-2": 11}

HEADER3_STANDARD = ["구분", "교과(군)", "세부과목", None, "기준학점", "운영학점",
                    "1학년", None, "2학년", None, "3학년", None,
                    "이수학점", "필수이수학점"]
# 미리내고: 과목유형/세부교과목 분리형 헤더(헤더 일반화 시연)
HEADER3_VARIANT = ["구분", "교과(군)", "과목유형", "세부교과목", "기준학점", "운영학점",
                   "1학년", None, "2학년", None, "3학년", None,
                   "이수학점", "필수이수학점"]
SEMESTER_LABEL_ROW = [None, None, None, None, None, None,
                      "1학기", "2학기", "1학기", "2학기", "1학기", "2학기", None, None]


def drow(section=None, group=None, subtype=None, name=None, base=None, run=None, **sems) -> list:
    row: list = [None] * 14
    row[0], row[1], row[2], row[3] = section, group, subtype, name
    row[4], row[5] = base, run
    for key, value in sems.items():
        row[SEM_COL[key.replace("_", "-")]] = value
    return row


def write_school_workbook(path: Path, title: str, header3: list, rows: list, merges: list) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "2026학년도 신입생 3개년 교육과정 편제표"

    grid = [
        [f"2026학년도 신입생 3개년 교육과정 편제표({title})"] + [None] * 13,
        [title] + [None] * 13,
        header3,
        [None] * 14,
        SEMESTER_LABEL_ROW,
        *rows,
    ]
    for r, rowdata in enumerate(grid, start=1):
        for c, value in enumerate(rowdata, start=1):
            if value is not None:
                ws.cell(row=r, column=c, value=value)
    for merge in merges:
        ws.merge_cells(merge)
    wb.save(path)


def build_hanbit() -> tuple[list, list]:
    """한빛고 — 전 오염 요소 총망라(병합 전파·분배, 택N, 다과목 셀, 각주,
    (온공) 접미, 정제 대상 오염값 '3'')."""
    rows = []

    def add(*args, **kwargs):
        rows.append(drow(*args, **kwargs))
        return 6 + len(rows) - 1  # 방금 추가한 행의 시트 행번호(1-based)

    # --- 학교지정 공통(1학년) ---
    add("학교지정", "국어", "공통", "공통국어1", base=4, run=4, **{"1_1": 4})
    add(None, None, "공통", "공통국어2", base=4, run=4, **{"1_2": 4})
    add(None, "수학", "공통", "공통수학1", base=4, run=4, **{"1_1": 4})
    add(None, None, "공통", "공통수학2", base=4, run=4, **{"1_2": 4})
    add(None, "영어", "공통", "공통영어1", base=4, run=4, **{"1_1": 4})
    add(None, None, "공통", "공통영어2", base=4, run=4, **{"1_2": 4})
    add(None, "사회", "공통", "통합사회1", base=4, run=4, **{"1_1": 4})
    add(None, None, "공통", "통합사회2", base=4, run=4, **{"1_2": 4})
    add(None, "과학", "공통", "통합과학1", base=4, run=4, **{"1_1": 4})
    add(None, None, "공통", "통합과학2", base=4, run=4, **{"1_2": 4})
    add(None, None, "공통", "과학탐구실험1", base=1, run=1, **{"1_1": 1})
    add(None, None, "공통", "과학탐구실험2", base=1, run=1, **{"1_2": 1})
    add(None, "한국사", "공통", "한국사1", base=3, run=3, **{"1_1": 3})
    add(None, None, "공통", "한국사2", base=3, run=3, **{"1_2": 3})

    # --- 학생선택 일반선택(2학년) ---
    add("학생선택", "국어", "일반", "화법과 언어", base=4, run=4, **{"2_1": 4})
    add(None, None, "일반", "독서와 작문", base=4, run=4, **{"2_2": 4})
    add(None, "수학", "일반", "대수", base=4, run=4, **{"2_1": 4})
    add(None, None, "일반", "확률과 통계", base=4, run=4, **{"2_2": 4})
    add(None, "영어", "일반", "영어Ⅰ", base=4, run=4, **{"2_1": 4})
    add(None, None, "일반", "영어Ⅱ", base=4, run=4, **{"2_2": 4})
    add(None, "사회", "일반", "사회와 문화", base=4, run=4, **{"2_1": 4})
    add(None, None, "일반", "경제", base=4, run=4, **{"2_2": 4})

    # --- 오염 ① 적층 학점 병합("4\n4\n4" 한 셀에 세 과목분 학점, 분배 복원) ---
    add(None, "과학", "진로", "지구과학", base=4, run=4, **{"2_2": "4\n4\n4"})
    add(None, None, "진로", "과학과제 연구", base=4, run=4)
    add(None, None, "진로", "역학과 에너지", base=4, run=4)

    # --- 오염 ② 학기값 세로 병합(L열=3-2, 단일 앵커값 전파 복원) ---
    add(None, "사회", "진로", "세계사", base=4, run=4, **{"3_2": 4})
    add(None, None, "진로", "동아시아 역사 기행", base=4, run=4)
    add(None, None, "진로", "한국지리 탐구", base=4, run=4)

    # --- 오염 ③ 다과목 셀(콤마 구분 한 셀에 세 과목) ---
    add(None, "과학", "진로", "물리학, 화학, 생명과학", base=4, run=4, **{"3_1": 4})

    # --- 오염 ④ 학기열 택N 마커 + 하위 선택군(택3) ---
    add(None, "정보", "진로", "인공지능 기초", base=4, run=4, **{"2_1": "택3"})
    add(None, None, "진로", "데이터 과학", base=4, run=4)
    add(None, None, "진로", "소프트웨어와 생활", base=4, run=4)
    add(None, None, None, "소계")  # 택N 그룹 경계(소계 skip 행)

    # --- 오염 ⑤ (온공) 접미 → 구분이 공동교육과정으로 승격 ---
    add(None, "정보", "진로", "인공지능 수학(온공)", base=3, run=3, **{"2_2": 3})

    # --- 오염 ⑥ 정제 대상 오염값 '3'(글리프 잔여, master에 없어 미매칭 1건) ---
    add(None, "수학", "진로", "경제 수학", base=3, run=3, **{"3_1": "3'"})

    # --- 오염 ⑦ 각주 마커(★) ---
    add(None, "교양", "진로", "심리학★", base=3, run=3, **{"3_2": 3})

    # --- 오염 ⑧ 하단 범례 행(본표 제외 대상) ---
    add(None, None, None, "★ 온라인 공동교육과정으로 인정되는 과목입니다")

    assert len(rows) == 37, f"한빛고 행수 불일치: {len(rows)}"
    assert rows[38 - 6][3] == "소계"
    assert rows[42 - 6][3].startswith("★")

    merges = [
        "A6:A19", "A20:A41",
        "B6:B7", "B8:B9", "B10:B11", "B12:B13", "B14:B17", "B18:B19",
        "B20:B21", "B22:B23", "B24:B25", "B26:B27",
        "B28:B30", "B31:B33", "B35:B37",
        "J28:J30",  # 적층 학점 병합(2-2)
        "L31:L33",  # 학기값 세로 병합(3-2)
    ]
    return rows, merges


def build_mirinae() -> tuple[list, list]:
    """미리내고 — 세부교과목 헤더 변형 + 택2 선택군/다과목 셀/각주 시연."""
    rows = []

    def add(*args, **kwargs):
        rows.append(drow(*args, **kwargs))
        return 6 + len(rows) - 1

    # --- 학교지정 공통(1학년) — 한빛고와 동일 공통 교육과정 ---
    add("학교지정", "국어", "공통", "공통국어1", base=4, run=4, **{"1_1": 4})
    add(None, None, "공통", "공통국어2", base=4, run=4, **{"1_2": 4})
    add(None, "수학", "공통", "공통수학1", base=4, run=4, **{"1_1": 4})
    add(None, None, "공통", "공통수학2", base=4, run=4, **{"1_2": 4})
    add(None, "영어", "공통", "공통영어1", base=4, run=4, **{"1_1": 4})
    add(None, None, "공통", "공통영어2", base=4, run=4, **{"1_2": 4})
    add(None, "사회", "공통", "통합사회1", base=4, run=4, **{"1_1": 4})
    add(None, None, "공통", "통합사회2", base=4, run=4, **{"1_2": 4})
    add(None, "과학", "공통", "통합과학1", base=4, run=4, **{"1_1": 4})
    add(None, None, "공통", "통합과학2", base=4, run=4, **{"1_2": 4})
    add(None, None, "공통", "과학탐구실험1", base=1, run=1, **{"1_1": 1})
    add(None, None, "공통", "과학탐구실험2", base=1, run=1, **{"1_2": 1})
    add(None, "한국사", "공통", "한국사1", base=3, run=3, **{"1_1": 3})
    add(None, None, "공통", "한국사2", base=3, run=3, **{"1_2": 3})

    # --- 학생선택 일반선택(2학년) ---
    add("학생선택", "수학", "일반", "대수", base=4, run=4, **{"2_1": 4})
    add(None, None, "일반", "확률과 통계", base=4, run=4, **{"2_2": 4})
    add(None, "영어", "일반", "영어Ⅰ", base=4, run=4, **{"2_1": 4})
    add(None, None, "일반", "영어Ⅱ", base=4, run=4, **{"2_2": 4})
    add(None, "사회", "일반", "사회와 문화", base=4, run=4, **{"2_1": 4})
    add(None, None, "일반", "경제", base=4, run=4, **{"2_2": 4})

    # --- 학생선택 진로선택(3학년, 단순 개설) ---
    add(None, "과학", "진로", "물리학", base=4, run=4, **{"3_1": 4})
    add(None, None, "진로", "지구과학", base=4, run=4, **{"3_2": 4})
    add(None, "교양", "진로", "심리학", base=3, run=3, **{"3_1": 3})
    add(None, None, "진로", "논리학", base=3, run=3, **{"3_2": 3})

    # --- 오염 ① 다과목 셀 ---
    add(None, "과학", "진로", "화학, 생명과학", base=4, run=4, **{"2_2": 4})

    # --- 오염 ② 학기열 택N 마커 + 하위 선택군(택2) ---
    add(None, "정보", "진로", "데이터 과학", base=4, run=4, **{"2_1": "택2"})
    add(None, None, "진로", "인공지능 기초", base=4, run=4)
    add(None, None, None, "소계")  # 택N 그룹 경계

    # --- 오염 ③ 각주 마커(★) ---
    add(None, "예술", "진로", "미술★", base=3, run=3, **{"3_1": 3})

    # --- 오염 ④ (온공) 접미 ---
    add(None, "사회", "진로", "한국지리 탐구(온공)", base=3, run=3, **{"3_2": 3})

    # --- 오염 ⑤ 하단 범례 행 ---
    add(None, None, None, "☆ 학교간 공동교육과정으로 개설되는 과목입니다")

    assert len(rows) == 31, f"미리내고 행수 불일치: {len(rows)}"
    assert rows[33 - 6][3] == "소계"
    assert rows[36 - 6][3].startswith("☆")

    merges = [
        "A6:A19", "A20:A35",
        "B6:B7", "B8:B9", "B10:B11", "B12:B13", "B14:B17", "B18:B19",
        "B20:B21", "B22:B23", "B24:B25",
        "B26:B27", "B28:B29",
        "B31:B32",
    ]
    return rows, merges


def main() -> int:
    SAMPLE_DIR.mkdir(parents=True, exist_ok=True)
    SOURCE_DIR.mkdir(parents=True, exist_ok=True)

    master_wb = build_master_workbook()
    master_path = SAMPLE_DIR / "master.xlsx"
    master_wb.save(master_path)
    print(f"저장: {master_path}")

    hanbit_rows, hanbit_merges = build_hanbit()
    hanbit_path = SOURCE_DIR / "한빛고 2026학년도 신입생 3개년 교육과정 편제표.xlsx"
    write_school_workbook(hanbit_path, "한빛고", HEADER3_STANDARD, hanbit_rows, hanbit_merges)
    print(f"저장: {hanbit_path}  ({len(hanbit_rows)}행)")

    mirinae_rows, mirinae_merges = build_mirinae()
    mirinae_path = SOURCE_DIR / "미리내고 2026학년도 신입생 3개년 교육과정 편제표.xlsx"
    write_school_workbook(mirinae_path, "미리내고", HEADER3_VARIANT, mirinae_rows, mirinae_merges)
    print(f"저장: {mirinae_path}  ({len(mirinae_rows)}행)")

    sheetmap_path = SAMPLE_DIR / "sheetmap.csv"
    with sheetmap_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["school", "file_keyword", "sheet_name", "sheet_confirmed", "note"])
        for rec in SHEETMAP_RECORDS:
            writer.writerow([rec["school"], rec["file_keyword"], rec["sheet_name"],
                              rec["sheet_confirmed"], rec["note"]])
    print(f"저장: {sheetmap_path}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
