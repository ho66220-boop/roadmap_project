"""Collect small, source-linked admission result samples from ADIGA.

This script only queries public ADIGA pages and filters results to universities
already present in the local roadmap workbook. Keep request limits conservative.

Usage:
    python scripts/collect_adiga_admission_results.py --max-queries 10 --max-popups 50
"""

from __future__ import annotations

import argparse
import csv
from dataclasses import dataclass
from datetime import datetime, timezone
from html import unescape
from html.parser import HTMLParser
import json
from pathlib import Path
import re
import sys
import time
from typing import Any
from urllib.parse import urlencode
from urllib.request import Request, urlopen

import openpyxl


BASE_URL = "https://www.adiga.kr"
DEFAULT_WORKBOOK = Path(__file__).resolve().parents[2] / "울산고교_과목로드맵_수정.xlsx"
DEFAULT_OUTPUT = Path(__file__).resolve().parents[1] / "data" / "admission_results_sample.csv"


@dataclass(frozen=True)
class SearchHit:
    university: str
    major: str
    unv_cd: str
    ru_cd: str
    search_year: str
    area: str
    recruit_count_hint: str
    competition_rate_hint: str


class TableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.in_tr = False
        self.in_cell = False
        self.current_cell: list[str] = []
        self.current_row: list[str] = []
        self.rows: list[list[str]] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag == "tr":
            self.in_tr = True
            self.current_row = []
        if self.in_tr and tag in {"td", "th"}:
            self.in_cell = True
            self.current_cell = []

    def handle_data(self, data: str) -> None:
        if self.in_cell:
            self.current_cell.append(data)

    def handle_endtag(self, tag: str) -> None:
        if self.in_tr and tag in {"td", "th"} and self.in_cell:
            text = normalize_space(" ".join(self.current_cell))
            self.current_row.append(text)
            self.in_cell = False
        if tag == "tr" and self.in_tr:
            if self.current_row:
                self.rows.append(self.current_row)
            self.in_tr = False


def normalize_space(text: str) -> str:
    return re.sub(r"\s+", " ", unescape(strip_tags(text or ""))).strip()


def strip_tags(text: str) -> str:
    return re.sub(r"<[^>]+>", "", text)


def normalize_name(text: str) -> str:
    text = normalize_space(text)
    text = re.sub(r"\[[^\]]+\]|\([^)]*\)", "", text)
    text = text.replace("대학교", "대")
    text = text.replace("국립", "")
    return re.sub(r"[^가-힣A-Za-z0-9]", "", text).lower()


def request_text(url: str, params: dict[str, Any] | None = None) -> str:
    full_url = url if not params else f"{url}?{urlencode(params)}"
    request = Request(
        full_url,
        headers={
            "User-Agent": "roadmap-project-admission-sample/0.1",
            "Referer": "https://www.adiga.kr/",
            "X-Requested-With": "XMLHttpRequest",
        },
    )
    with urlopen(request, timeout=20) as response:
        return response.read().decode("utf-8", errors="replace")


def workbook_universities(path: Path) -> set[str]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["대학트랙"]
    return {
        normalize_name(row[0])
        for row in sheet.iter_rows(min_row=2, values_only=True)
        if row and row[0]
    }


def workbook_major_terms(path: Path) -> list[str]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "로드맵DB_v2_대학트랙집계" in workbook.sheetnames:
        sheet = workbook["로드맵DB_v2_대학트랙집계"]
        majors = [
            normalize_space(row[1])
            for row in sheet.iter_rows(min_row=2, values_only=True)
            if row and len(row) > 1 and row[1]
        ]
    else:
        sheet = workbook["대학트랙"]
        majors = [
            normalize_space(row[1])
            for row in sheet.iter_rows(min_row=2, values_only=True)
            if row and len(row) > 1 and row[1]
        ]
    seen = set()
    result = []
    for major in majors:
        key = normalize_name(major)
        if key and key not in seen:
            seen.add(key)
            result.append(major)
    return result


def search_major(term: str, limit: int) -> list[SearchHit]:
    payload = request_text(
        f"{BASE_URL}/man/sch/majorInfo2.do",
        {
            "offset": 0,
            "limit": limit,
            "search": term,
            "sort": "$relevance",
            "sortType": "desc",
            "reSearch": "",
            "exceptSearch": "",
            "sortYear": "",
            "accuracySearch": "",
            "checkMiss": "",
        },
    )
    if not payload or payload.strip() == "null":
        return []
    data = json.loads(payload)
    hits = []
    for row in data.get("result", {}).get("rows", []):
        fields = row.get("fields", {})
        unv_cd = normalize_space(fields.get("UNV_CD", ""))
        ru_cd = normalize_space(fields.get("RCU_CD") or fields.get("RU_CD") or fields.get("MAPNG_GRP_RCU_CD") or "")
        if not unv_cd or not ru_cd:
            continue
        hits.append(
            SearchHit(
                university=normalize_space(fields.get("UNIV_ALL_NM_SORT") or fields.get("UNIV_ALL_NM") or ""),
                major=normalize_space(fields.get("CMMN_SUBJCT_NM_SORT") or fields.get("CMMN_SUBJCT_NM") or ""),
                unv_cd=unv_cd,
                ru_cd=ru_cd,
                search_year=normalize_space(fields.get("SYR", "")),
                area=normalize_space(fields.get("AREA_NM", "")),
                recruit_count_hint=normalize_space(fields.get("ESPCPA_INWON", "")),
                competition_rate_hint=normalize_space(fields.get("ATRCR_CMPET_RT", "")),
            )
        )
    return hits


def parse_admission_popup(hit: SearchHit, result_year: str) -> list[dict[str, str]]:
    url = f"{BASE_URL}/ucp/cls/uni/classUnivAdmssPopup.do"
    source_url = f"{url}?{urlencode({'searchSyr': result_year, 'unvCd': hit.unv_cd, 'ruCd': hit.ru_cd})}"
    html = request_text(url, {"searchSyr": result_year, "unvCd": hit.unv_cd, "ruCd": hit.ru_cd})
    parser = TableParser()
    parser.feed(html)

    rows = []
    for cells in parser.rows:
        if not cells or cells[0] not in {"수시", "정시(가)", "정시(나)", "정시(다)", "추가"}:
            continue
        if len(cells) < 10:
            continue
        score_50 = parse_cut_value(cells[6] if len(cells) > 6 else "", "50")
        score_70 = parse_cut_value(cells[7] if len(cells) > 7 else "", "70")
        grade_50 = parse_cut_value(cells[8] if len(cells) > 8 else "", "50")
        grade_70 = parse_cut_value(cells[9] if len(cells) > 9 else "", "70")
        rows.append(
            {
                "result_year": result_year,
                "university": hit.university,
                "major": hit.major,
                "area": hit.area,
                "admission_period": cells[0],
                "admission_type": cells[1] if len(cells) > 1 else "",
                "track_name": cells[2] if len(cells) > 2 else "",
                "recruit_count": cells[3] if len(cells) > 3 else "",
                "additional_pass_count": cells[4] if len(cells) > 4 else "",
                "competition_rate": cells[5] if len(cells) > 5 else "",
                "score_50": score_50,
                "score_70": score_70,
                "grade_50": grade_50,
                "grade_70": grade_70,
                "source_url": source_url,
                "collected_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            }
        )
    return rows


def parse_cut_value(text: str, cut: str) -> str:
    clean = normalize_space(text)
    match = re.search(rf"{re.escape(cut)}%\(([^)]*)\)", clean)
    return match.group(1).strip() if match else ""


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "result_year",
        "university",
        "major",
        "area",
        "admission_period",
        "admission_type",
        "track_name",
        "recruit_count",
        "additional_pass_count",
        "competition_rate",
        "score_50",
        "score_70",
        "grade_50",
        "grade_70",
        "source_url",
        "collected_at",
    ]
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--result-year", default="2025")
    parser.add_argument("--max-queries", type=int, default=10)
    parser.add_argument("--search-limit", type=int, default=50)
    parser.add_argument("--max-popups", type=int, default=50)
    parser.add_argument("--delay", type=float, default=0.8)
    parser.add_argument(
        "--terms",
        default="",
        help="Comma-separated major search terms. Defaults to workbook major terms.",
    )
    args = parser.parse_args()

    target_universities = workbook_universities(args.workbook)
    if args.terms:
        terms = [normalize_space(term) for term in args.terms.split(",") if normalize_space(term)]
    else:
        terms = workbook_major_terms(args.workbook)[: args.max_queries]
    seen_popups: set[tuple[str, str]] = set()
    collected: list[dict[str, str]] = []
    popup_count = 0

    for term in terms:
        print(f"[search] {term}")
        try:
            hits = search_major(term, args.search_limit)
        except Exception as exc:
            print(f"  ! search failed: {exc}")
            continue
        time.sleep(args.delay)

        for hit in hits:
            if normalize_name(hit.university) not in target_universities:
                continue
            popup_key = (hit.unv_cd, hit.ru_cd)
            if popup_key in seen_popups:
                continue
            if popup_count >= args.max_popups:
                write_csv(args.output, collected)
                print(f"[done] rows={len(collected)} popups={popup_count} output={args.output}")
                return 0
            seen_popups.add(popup_key)
            popup_count += 1
            print(f"  [popup] {hit.university} / {hit.major}")
            try:
                collected.extend(parse_admission_popup(hit, args.result_year))
            except Exception as exc:
                print(f"    ! popup failed: {exc}")
            time.sleep(args.delay)

    write_csv(args.output, collected)
    print(f"[done] rows={len(collected)} popups={popup_count} output={args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
