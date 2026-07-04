# -*- coding: utf-8 -*-
"""편제표 파서 v2 테스트 (unittest, 기존 컨벤션).

합성 워크북은 make_ws(rows, merges) 로 인메모리 생성한다(임시 파일 불필요).
데이터는 시트 5행(0-based idx 4)부터. 학기열(0-based): 1-1=6,1-2=7,2-1=8,2-2=9,3-1=10,3-2=11.
"""
from __future__ import annotations

import sys
import unicodedata
import unittest
from pathlib import Path

_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(_ROOT))
sys.path.insert(0, str(_ROOT / "scripts"))

from openpyxl import Workbook

from curriculum_parser import parse_workbook
from curriculum_parser import cells
from curriculum_parser.cells import clean_value, norm
from curriculum_parser.header import build_column_map, find_header
from curriculum_parser.loader import resolve_merge_semantics

# --- 공통 헤더(구분/교과/세부과목/name/기준/운영/1~3학년/이수/필수) ---
HEADER = [
    ["2026학년도 신입생 3개년 교육과정 편제표(테스트)"],
    ["테스트고등학교"],
    ["구분", "교과(군)", "세부과목", None, "기준학점", "운영학점",
     "1학년", None, "2학년", None, "3학년", None, "이수학점", "필수이수학점"],
    [None, None, None, None, None, None,
     "1학기", "2학기", "1학기", "2학기", "1학기", "2학기", None, None],
]
SEM_COL = {"1-1": 6, "1-2": 7, "2-1": 8, "2-2": 9, "3-1": 10, "3-2": 11}

MASTER = {norm(n): n for n in [
    "문학", "대수", "영어Ⅰ", "물리학", "화학", "생명과학", "지구과학",
    "중국어", "일본어", "음악", "논리와 사고", "세계사", "사회와 문화",
    "확률과 통계", "미적분Ⅰ", "미적분Ⅱ", "정보", "전자기와 양자", "미술",
]}


def make_ws(rows, merges=()):
    """rows: list[list](0-based 열). 병합의 비앵커 셀엔 값을 넣지 말 것(실파일과 동일)."""
    wb = Workbook()
    ws = wb.active
    for r, rowdata in enumerate(rows, start=1):
        for c, val in enumerate(rowdata, start=1):
            if val is not None:
                ws.cell(row=r, column=c, value=val)
    for m in merges:
        ws.merge_cells(m)
    return ws


def drow(section=None, group=None, subtype=None, name=None,
         base=4, run=4, **sems):
    row = [section, group, subtype, name, base, run,
           None, None, None, None, None, None, None, None]
    for k, v in sems.items():
        row[SEM_COL[k.replace("_", "-")]] = v
    return row


def run_parse(data_rows, merges=()):
    ws = make_ws(HEADER + data_rows, merges)
    return parse_workbook(ws, "테스트고", "test.xlsx", MASTER)


def by_name(result):
    out = {}
    for r in result.rows:
        out.setdefault(cells.strip_markers(r.raw_name), r)
    return out


class HeaderTest(unittest.TestCase):
    def test_find_header_and_columns(self):
        grid = [list(r) + [None] * (14 - len(r)) for r in HEADER]
        idx = find_header(grid)
        self.assertEqual(idx, 2)
        col = build_column_map(grid, idx)
        self.assertIsNotNone(col)
        self.assertEqual(col["name"], 3)
        self.assertEqual(col["base"], 4)
        self.assertEqual(col["run"], 5)
        self.assertEqual(col["semesters"], SEM_COL)
        self.assertEqual(col["data_start"], 4)

    def test_separate_type_and_sebugyogwamok_name(self):
        # 대송/무거/우신/신선여고형: 과목유형(col2) + 세부교과목(col3=과목명)
        grid = [
            ["제목"],
            [None],
            ["구분", "교과(군)", "과목유형", "세부교과목", "기준학점", "운영학점",
             "1학년", None, "2학년", None, "3학년", None],
            [None, None, None, None, None, None,
             "1학기", "2학기", "1학기", "2학기", "1학기", "2학기"],
        ]
        grid = [list(r) + [None] * (12 - len(r)) for r in grid]
        idx = find_header(grid)
        self.assertEqual(idx, 2)
        col = build_column_map(grid, idx)
        self.assertIsNotNone(col)
        self.assertEqual(col["type"], 2)
        self.assertEqual(col["name"], 3)      # 세부교과목이 과목명 열
        self.assertEqual(col["semesters"]["2-1"], 8)

    def test_header_without_gubun_and_sebugwamok(self):
        # 울산고운고형: 구분/세부과목 없음, [_,교과,과목,운영학점,1학년...]
        grid = [
            ["울산고운고 2026학년도"],
            [None],
            [None, "교과", "과목", "운영학점", "1학년", None, "2학년", None,
             "3학년", None, "이수학점"],
            [None, None, None, None, "1학기", "2학기", "1학기", "2학기",
             "1학기", "2학기", None],
        ]
        grid = [list(r) + [None] * (11 - len(r)) for r in grid]
        idx = find_header(grid)
        self.assertEqual(idx, 2)             # 교과+과목+학년으로 탐지
        col = build_column_map(grid, idx)
        self.assertIsNotNone(col)
        self.assertEqual(col["group"], 1)
        self.assertEqual(col["name"], 2)
        self.assertEqual(col["run"], 3)
        self.assertEqual(col["semesters"],
                         {"1-1": 4, "1-2": 5, "2-1": 6, "2-2": 7, "3-1": 8, "3-2": 9})


class MergeLoaderTest(unittest.TestCase):
    def test_propagate(self):
        rows = [
            drow("학교지정", "국어", "공통", "문학"),
            drow(None, None, "공통", "대수"),
        ]
        # 1-1 열(sheet col G=7) 을 5~6행 세로병합, 앵커 4
        rows[0][SEM_COL["1-1"]] = 4
        res = run_parse(rows, merges=["G5:G6"])
        idx = by_name(res)
        self.assertEqual(idx["문학"].credits.get("1-1"), 4)
        self.assertEqual(idx["대수"].credits.get("1-1"), 4)

    def test_distribute_stacked(self):
        # '4\n4\n4' 를 3행 세로병합 → 행별 4 분배
        rows = [
            drow("학교지정", "과학", "일반", "물리학"),
            drow(None, None, "일반", "화학"),
            drow(None, None, "일반", "생명과학"),
        ]
        rows[0][SEM_COL["2-1"]] = "4\n4\n4"     # I5, 병합 I5:I7
        res = run_parse(rows, merges=["I5:I7"])
        idx = by_name(res)
        for nm in ("물리학", "화학", "생명과학"):
            self.assertEqual(idx[nm].credits.get("2-1"), 4, nm)
        self.assertFalse([lg for lg in res.logs if lg.flag == "stack_mismatch"])

    def test_distribute_stack_mismatch(self):
        # '4\n4\n4'(3토큰) 를 5행 병합 → 부족분 마지막값 반복 + stack_mismatch
        rows = [
            drow("학교지정", "과학", "일반", f"과목{i}") for i in range(5)
        ]
        rows[0][SEM_COL["2-1"]] = "4\n4\n4"
        res = run_parse(rows, merges=["I5:I9"])
        idx = by_name(res)
        for i in range(5):
            self.assertEqual(idx[f"과목{i}"].credits.get("2-1"), 4)
        self.assertTrue([lg for lg in res.logs if lg.flag == "stack_mismatch"])


class ChoiceGroupTest(unittest.TestCase):
    def test_semester_column_marker(self):
        # 마커 '택3' (2-1) + 멤버 3과목(분배된 4)
        rows = [
            drow("학생선택", "국어", "일반", "문학"),
            drow(None, "수학", "일반", "대수"),
            drow(None, "영어", "일반", "영어Ⅰ"),
            drow(None, "과학", "일반", "물리학"),
        ]
        rows[0][SEM_COL["2-1"]] = "택3"
        rows[1][SEM_COL["2-1"]] = "4\n4\n4"   # I6:I8 분배 → 대수/영어Ⅰ/물리학
        res = run_parse(rows, merges=["I6:I8"])
        idx = by_name(res)
        cid = idx["문학"].choice_id
        self.assertTrue(cid)
        self.assertTrue(cid.startswith("테스트고-2-1-"))
        for nm in ("문학", "대수", "영어Ⅰ", "물리학"):
            self.assertEqual(idx[nm].choice_id, cid, nm)
            self.assertEqual(idx[nm].take_n, 3, nm)
            self.assertGreater(idx[nm].credits.get("2-1", 0), 0, nm)

    def test_name_suffix_marker(self):
        # 과목명 접미 (택1)
        rows = [drow("학생선택", "제2외국어", "일반", "중국어, 일본어(택1)")]
        rows[0][SEM_COL["2-1"]] = 3
        res = run_parse(rows)
        idx = by_name(res)
        self.assertIn("중국어", idx)
        self.assertIn("일본어", idx)
        self.assertEqual(idx["중국어"].take_n, 1)
        self.assertEqual(idx["중국어"].choice_id, idx["일본어"].choice_id)
        self.assertEqual(idx["중국어"].official_name, "중국어")  # 접미 제거 후 매칭

    def test_other_column_marker(self):
        # 기타열(이수학점 col idx12)에 택2, 멤버는 자기 학기 유지
        rows = [drow("학생선택", "사회", "일반", "세계사, 사회와 문화")]
        rows[0][SEM_COL["3-1"]] = 4
        rows[0][12] = "택2"   # 이수학점 열
        res = run_parse(rows)
        idx = by_name(res)
        self.assertEqual(idx["세계사"].take_n, 2)
        self.assertTrue(idx["세계사"].choice_id)
        self.assertEqual(idx["세계사"].credits.get("3-1"), 4)


class CrossTest(unittest.TestCase):
    def test_cross_pair_opens_both(self):
        rows = [drow("학생선택", "예술", "진로", "음악↔논리와 사고")]
        rows[0][SEM_COL["3-1"]] = 3
        res = run_parse(rows)
        idx = by_name(res)
        self.assertIn("음악", idx)
        self.assertIn("논리와 사고", idx)
        self.assertTrue(idx["음악"].cross)
        self.assertTrue(idx["논리와 사고"].cross)
        self.assertEqual(idx["음악"].credits.get("3-1"), 3)
        self.assertEqual(idx["논리와 사고"].credits.get("3-1"), 3)


class CleanValueTest(unittest.TestCase):
    def test_prime_glyph(self):
        self.assertEqual(clean_value("3'")[0], 3)
        self.assertEqual(clean_value("3'")[1], "glyph")

    def test_degree_glyph(self):
        self.assertEqual(clean_value("3º")[0], 3)
        self.assertEqual(clean_value("3º")[1], "glyph")

    def test_range_value(self):
        self.assertEqual(clean_value("29~30"), (29, "range_value"))

    def test_paren_note(self):
        self.assertEqual(clean_value("16\n(70)"), (16, "paren_note"))

    def test_multi_credit(self):
        self.assertEqual(clean_value("3\n3(4)"), (3, "multi_credit"))

    def test_plain(self):
        self.assertEqual(clean_value("4"), (4, ""))
        self.assertEqual(clean_value(None), (None, ""))


class CellSemesterCleanTest(unittest.TestCase):
    def test_range_in_semester_cell_logs(self):
        rows = [drow("학교지정", "국어", "공통", "문학")]
        rows[0][SEM_COL["1-1"]] = "29~30"
        res = run_parse(rows)
        idx = by_name(res)
        self.assertEqual(idx["문학"].credits.get("1-1"), 29)
        self.assertTrue([lg for lg in res.logs if lg.flag == "range_value"])


class LegendMultiMarkerTest(unittest.TestCase):
    def test_legend_row_skipped(self):
        rows = [
            drow("학교지정", "국어", "공통", "문학", **{"1-1": 4}),
            drow(None, None, None, "☆ 고시외 과목"),
        ]
        res = run_parse(rows)
        idx = by_name(res)
        self.assertIn("문학", idx)
        self.assertNotIn("고시외 과목", idx)

    def test_multi_subject_split(self):
        rows = [drow("학생선택", "과학", "일반", "물리학, 화학, 생명과학", **{"2-1": 4})]
        res = run_parse(rows)
        names = {cells.strip_markers(r.raw_name) for r in res.rows}
        self.assertSetEqual(names, {"물리학", "화학", "생명과학"})

    def test_marker_stripped_and_matched(self):
        rows = [drow("학생선택", "과학", "일반", "물리학■", **{"2-1": 3})]
        res = run_parse(rows)
        idx = by_name(res)
        self.assertEqual(idx["물리학"].official_name, "물리학")

    def test_small_square_marker_stripped(self):
        # 성신고형 ▪(U+25AA) 접미 마커 — strip 후 매칭돼야 함
        rows = [drow("학생선택", "과학", "일반", "생명과학▪", **{"2-1": 3})]
        res = run_parse(rows)
        idx = by_name(res)
        self.assertEqual(idx["생명과학"].official_name, "생명과학")

    def test_small_square_multi_subject_split(self):
        # 성신고형 '물리학▪ 화학▪ 생명과학▪' — ▪가 구분자 겸 마커
        rows = [drow("학생선택", "과학", "진로", "물리학▪ 화학▪ 생명과학▪", **{"2-2": 4})]
        res = run_parse(rows)
        names = {cells.strip_markers(r.raw_name) for r in res.rows}
        self.assertSetEqual(names, {"물리학", "화학", "생명과학"})
        for r in res.rows:
            self.assertTrue(r.official_name, r.raw_name)


class NfdTest(unittest.TestCase):
    def test_nfd_name_matches(self):
        nfd_name = unicodedata.normalize("NFD", "물리학")
        self.assertNotEqual(nfd_name, "물리학")  # 분해형
        rows = [drow("학생선택", "과학", "일반", nfd_name, **{"2-1": 3})]
        res = run_parse(rows)
        matched = [r for r in res.rows if r.official_name == "물리학"]
        self.assertTrue(matched)


class MetaNoteTest(unittest.TestCase):
    def test_bottom_note_is_meta(self):
        note = ("공동교육과정 이수 시기별 2학점 또는 3학점으로 운영할 수 있다. "
                "희망하는 학생은 별도로 계획하여 운영한다.")
        rows = [
            drow("학생선택", "국어", "일반", "문학", **{"1-1": 4}),
            [None, None, None, note, None, None, None, None, None, None, None, None, None, None],
        ]
        res = run_parse(rows)
        names = {cells.strip_markers(r.raw_name) for r in res.rows}
        self.assertNotIn(cells.strip_markers(note), names)
        self.assertTrue(any("운영할 수 있다" in m for m in res.meta_notes))


class JointProgramSuffixTest(unittest.TestCase):
    def test_joint_suffix_promotes_section_and_strips_from_name(self):
        # 변경 3: '(온공)' 접미 -> 매칭 전 제거 + 해당 행 구분을 공동교육과정으로 승격
        rows = [drow("학생선택", "과학", "일반", "물리학(온공)", **{"2-1": 3})]
        res = run_parse(rows)
        self.assertEqual(len(res.rows), 1)
        row = res.rows[0]
        self.assertEqual(row.official_name, "물리학")
        self.assertEqual(row.section, "공동교육과정")

    def test_non_joint_row_keeps_original_section(self):
        rows = [drow("학생선택", "과학", "일반", "물리학", **{"2-1": 3})]
        res = run_parse(rows)
        self.assertEqual(res.rows[0].section, "학생선택")

    def test_joint_suffix_with_trailing_footnote_marker_still_promotes(self):
        # 약사고형: '(온공)' 뒤에 각주 마커(▲)가 덧붙는 경우도 승격돼야 함
        rows = [drow("학생선택", "정보", "진로", "프로그래밍★(온공)▲", **{"2-1": 3})]
        res = run_parse(rows)
        self.assertEqual(res.rows[0].section, "공동교육과정")


class HalfWidthLatinAliasTest(unittest.TestCase):
    def test_half_width_misugbun_one_maps_to_roman_numeral(self):
        rows = [drow("학생선택", "수학", "진로", "미적분I", **{"2-1": 3})]
        res = run_parse(rows)
        self.assertEqual(res.rows[0].official_name, "미적분Ⅰ")

    def test_half_width_misugbun_two_maps_to_roman_numeral(self):
        rows = [drow("학생선택", "수학", "진로", "미적분II", **{"2-1": 3})]
        res = run_parse(rows)
        self.assertEqual(res.rows[0].official_name, "미적분Ⅱ")

    def test_typo_jeongi_maps_to_jeonja(self):
        rows = [drow("학생선택", "과학", "진로", "전지기와 양자", **{"2-1": 3})]
        res = run_parse(rows)
        self.assertEqual(res.rows[0].official_name, "전자기와 양자")

    def test_misul_school_designated_suffix_maps_to_misul(self):
        rows = [drow("학교지정", "예술", "공통", "미술(학교 지정)", **{"1-1": 3})]
        res = run_parse(rows)
        self.assertEqual(res.rows[0].official_name, "미술")


class CompatTest(unittest.TestCase):
    def test_output_loads_into_engine(self):
        from roadmap_engine.recommender import DataStore, RecommendationEngine
        from curriculum_parser.report import build_workbook
        rows = [
            drow("학교지정", "국어", "공통", "문학", **{"1-1": 4}),
            drow("학생선택", "과학", "일반", "물리학, 화학", **{"2-1": 4}),
        ]
        res = run_parse(rows)
        wb = build_workbook(res.rows, [{
            "school": "테스트고", "file": "test.xlsx", "sheet": "s",
            "rows": len(res.rows), "matched": 0, "sem_cov": 0, "groups": 0,
            "sheet_confirmed": "Y", "issues": [], "logs": [], "meta_notes": [],
            "group_objs": [],
        }])
        tmp = Path(__file__).resolve().parent / "_tmp_compat.xlsx"
        wb.save(tmp)
        try:
            engine = RecommendationEngine.__new__(RecommendationEngine)
            engine.store = DataStore()
            engine.curriculum_path = tmp
            engine._load_curriculum(engine.store)
            self.assertIn("테스트고", engine.store.schools)
            subs = {o.subject for o in engine.store.schools["테스트고"]}
            self.assertIn("물리학", subs)
        finally:
            tmp.unlink(missing_ok=True)


if __name__ == "__main__":
    unittest.main()
