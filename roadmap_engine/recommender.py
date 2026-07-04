# -*- coding: utf-8 -*-
"""3-layer 과목 추천 엔진.

데이터 소스
  1. 로드맵 엑셀(울산고교_과목로드맵_수정.xlsx)
     - 학과추천: 커리어넷 기반 학과별 일반/진로/융합 추천 과목  → Layer 1
     - 대학트랙: 대교협 기반 대학·모집단위별 핵심/권장 과목      → Layer 2
     - 대학제시율_학과/계열: 과목별 제시 비율(추천 근거·정렬)
     - 과목마스터: 공식 과목명 정규화 키
  2. 전처리 편제표(data/curriculum_g1_2026.xlsx, scripts/build_curriculum.py 생성)
     - 학교별 고1(2026 입학) 편제표                             → Layer 3

추천 흐름: 목표 학과의 포괄 추천(L1) + 대학별 추가 추천(L2)을 합쳐
학생 학교의 편제표와 대조(L3)해 "개설 과목/학기"와 "미개설 과목"으로 나눈다.
"""
from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
import re
from typing import Any

import openpyxl


SEMESTER_KEYS = ["1-1", "1-2", "2-1", "2-2", "3-1", "3-2"]

# 대학트랙 과목명에 등장하는 교과(군) 단위 포괄 표현 — 개별 과목이 아니므로 편제표 대조에서 제외
GROUP_TERMS = {
    "국어", "수학", "영어", "사회", "과학", "한국사", "체육", "예술",
    "기술가정", "정보", "제2외국어", "한문", "교양", "사회역사도덕포함",
    # 구 교육과정식 대학트랙 표현(2022 개정 개별 과목이 아닌 교과 단위 포괄)
    "역사", "윤리", "지리", "일반사회", "과학교과", "수학1", "수학2",
    "한국사12", "전과목", "과학전과목", "수학전과목",
    "제2외국어관련과목", "제2외국어과목", "제3외국어과목",
    # 원문이 원문자(circled digit) 글리프를 써서 norm()이 로마숫자처럼 변환하지 못하는 실측 표기
    "수학①", "수학②",
}

# 대학트랙의 구식/포괄 과목 표기 -> 2022 개정 공식 과목명
L2_SUBJECT_ALIASES = {
    "미적분": ["미적분Ⅰ", "미적분Ⅱ"],
    "물리": ["물리학"],
    "한국지리": ["한국지리 탐구"],
    "동아시아사": ["동아시아 역사 기행"],
    "수학(특히 미적분)": ["미적분Ⅰ", "미적분Ⅱ"],
    "프랑스 회화": ["프랑스어 회화"],
    "도시와 미래탐구": ["도시의 미래 탐구"],
    "과학과제 탐구": ["과학과제 연구"],
    "물리과 에너지": ["역학과 에너지"],
    "독해와 작문": ["영어 독해와 작문"],
}
SEMESTER_LABELS = {
    "1-1": "1학년 1학기", "1-2": "1학년 2학기",
    "2-1": "2학년 1학기", "2-2": "2학년 2학기",
    "3-1": "3학년 1학기", "3-2": "3학년 2학기",
}


def repair_text(value: Any) -> str:
    """UTF-8 텍스트가 latin-1/cp1252로 잘못 디코딩된 경우 복구한다."""
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    if any(marker in text for marker in ("ì", "í", "ê", "ë", "â")):
        for encoding in ("latin1", "cp1252"):
            try:
                repaired = text.encode(encoding).decode("utf-8")
            except UnicodeError:
                continue
            if repaired and repaired != text:
                return repaired.strip()
    return text


def norm(value: Any) -> str:
    text = repair_text(value).lower()
    text = text.replace("Ⅰ", "1").replace("Ⅱ", "2")
    text = re.sub(r"[\s·ㆍ․,/_()\-\[\]]+", "", text)
    text = text.replace("대학교", "대")
    return text


def to_float(value: Any, default: float = 0.0) -> float:
    try:
        if value in (None, "", "-"):
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def split_subjects(text: str) -> list[str]:
    raw = re.split(r"[,/|]+|\n", text or "")
    return [repair_text(item).strip() for item in raw if repair_text(item).strip()]


@dataclass(frozen=True)
class Offering:
    """학교 편제표의 과목 한 줄."""
    subject: str          # 공식과목명(매칭 실패 시 원본명)
    matched: bool
    section: str          # 학교지정/학생선택 등
    group: str            # 교과군
    subject_type: str     # 공통/일반/진로/융합
    semesters: tuple[str, ...]
    choice_group: str = ""  # 선택군ID(택N 그룹). 미해당 시 ""
    take_n: int = 0          # 택N의 N. 미해당 시 0


@dataclass(frozen=True)
class UniTrackRow:
    university: str
    unit: str             # 모집단위
    subject: str
    track_type: str       # 핵심/권장
    priority: float
    comment: str


@dataclass
class DataStore:
    schools: dict[str, list[Offering]] = field(default_factory=dict)
    all_offered_keys: set = field(default_factory=set)                          # 전 학교 개설과목 norm키(공통 미개설 판정용)
    dept_recs: dict[str, dict[str, list[str]]] = field(default_factory=dict)   # 학과 -> 추천구분 -> 과목들
    dept_track: dict[str, str] = field(default_factory=dict)                    # 학과 -> 계열
    uni_tracks: list[UniTrackRow] = field(default_factory=list)
    rate_by_dept: dict[str, dict[str, float]] = field(default_factory=dict)     # 대상학과 -> 과목 -> 비율%
    rate_by_track: dict[str, dict[str, float]] = field(default_factory=dict)    # 계열 -> 과목 -> 비율%
    subject_master: dict[str, str] = field(default_factory=dict)                # 정규화키 -> 공식과목명


class RecommendationEngine:
    def __init__(self, workbook_path: Path, curriculum_path: Path | None = None):
        project_root = Path(__file__).resolve().parents[1]
        self.workbook_path = Path(workbook_path)
        self.curriculum_path = Path(curriculum_path) if curriculum_path else project_root / "data" / "curriculum_g1_2026.xlsx"
        self.store = self._load()

    # ------------------------------------------------------------------ load
    def _load(self) -> DataStore:
        if not self.workbook_path.exists():
            raise FileNotFoundError(f"로드맵 엑셀 파일을 찾을 수 없습니다: {self.workbook_path}")
        if not self.curriculum_path.exists():
            raise FileNotFoundError(
                f"전처리 편제표가 없습니다: {self.curriculum_path}\n"
                "scripts/build_curriculum.py를 먼저 실행하세요."
            )
        store = DataStore()
        wb = openpyxl.load_workbook(self.workbook_path, read_only=True, data_only=True)
        self._load_master(wb, store)
        self._load_dept_recs(wb, store)
        self._load_uni_tracks(wb, store)
        self._load_rates(wb, store)
        wb.close()
        self._load_curriculum(store)
        return store

    def _load_master(self, wb, store: DataStore) -> None:
        for row in wb["과목마스터"].iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            official = repair_text(row[0])
            key = repair_text(row[1]) if len(row) > 1 and row[1] else norm(official)
            store.subject_master[key] = official
            store.subject_master[norm(official)] = official

    def _load_dept_recs(self, wb, store: DataStore) -> None:
        for row in wb["학과추천"].iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 4 or not row[0] or not row[3]:
                continue
            dept, track, kind, subject = (repair_text(c) for c in row[:4])
            store.dept_recs.setdefault(dept, {}).setdefault(kind or "일반", []).append(subject)
            if track:
                store.dept_track[dept] = track

    def _load_uni_tracks(self, wb, store: DataStore) -> None:
        for row in wb["대학트랙"].iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 4 or not row[0] or not row[2]:
                continue
            store.uni_tracks.append(
                UniTrackRow(
                    university=repair_text(row[0]),
                    unit=repair_text(row[1]),
                    subject=repair_text(row[2]),
                    track_type=repair_text(row[3]) or "권장",
                    priority=to_float(row[4], 99.0),
                    comment=repair_text(row[5]) if len(row) > 5 else "",
                )
            )

    def _load_rates(self, wb, store: DataStore) -> None:
        for sheet, target in (("대학제시율_학과", store.rate_by_dept), ("대학제시율_계열", store.rate_by_track)):
            if sheet not in wb.sheetnames:
                continue
            for row in wb[sheet].iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 5 or not row[0] or not row[1]:
                    continue
                target.setdefault(repair_text(row[0]), {})[repair_text(row[1])] = to_float(row[4])

    def _load_curriculum(self, store: DataStore) -> None:
        wb = openpyxl.load_workbook(self.curriculum_path, read_only=True, data_only=True)
        ws = wb["고1편제표"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            school = repair_text(row[0])
            semesters = tuple(key for i, key in enumerate(SEMESTER_KEYS) if to_float(row[10 + i]) > 0)
            subject = repair_text(row[6])
            offering = Offering(
                subject=subject,
                matched=repair_text(row[7]) == "O",
                section=repair_text(row[2]),
                group=repair_text(row[3]),
                subject_type=repair_text(row[4]),
                semesters=semesters,
                choice_group=repair_text(row[18]) if len(row) > 18 else "",
                take_n=int(to_float(row[19])) if len(row) > 19 else 0,
            )
            store.schools.setdefault(school, []).append(offering)
            store.all_offered_keys.add(norm(subject))
        wb.close()

    # ------------------------------------------------------------------ api
    def meta(self) -> dict[str, Any]:
        return {
            "workbook": self.workbook_path.name,
            "curriculum": self.curriculum_path.name,
            "school_count": len(self.store.schools),
            "dept_count": len(self.store.dept_recs),
            "university_count": len({t.university for t in self.store.uni_tracks}),
            "schools": sorted(self.store.schools),
            "majors": sorted(self.store.dept_recs)[:600],
        }

    def chat(self, message: str, profile: dict[str, Any] | None = None) -> dict[str, Any]:
        profile = profile or {}
        school = self._pick_school(profile.get("school") or message)
        major = self._pick_major(profile.get("major") or message)

        if not school:
            return self._need_more(
                "학교명을 알려주면 해당 학교 편제표 안에서 실제 선택 가능한 과목만 골라볼 수 있어요. "
                f"(지원 학교: {', '.join(sorted(self.store.schools))})"
            )
        if not major:
            return self._need_more("목표 학과나 계열을 알려주면 커리어넷·대교협 기준 추천 과목을 찾아볼게요.")

        result = self.recommend(school=school, major=major)
        result["answer"] = self._compose_answer(result)
        return result

    def recommend(self, school: str, major: str) -> dict[str, Any]:
        track = self.store.dept_track.get(major, "")
        layer1 = self._layer1_dept(major, track)
        layer2 = self._layer2_universities(major)
        layer3 = self._layer3_school(school, layer1, layer2)

        sources = [
            f"{self.workbook_path.name} / 학과추천(커리어넷)",
            f"{self.workbook_path.name} / 대학트랙(대교협 시행계획)",
            f"{self.workbook_path.name} / 대학제시율",
            f"{self.curriculum_path.name} / {school} 2026 신입생 3개년 편제표",
        ]
        return {
            "mode": "recommendation",
            "school": school,
            "major": major,
            "track": track,
            "layer1": layer1,
            "layer2": layer2,
            "layer3": layer3,
            "graph": self._graph_payload(school, major, layer3),
            "sources": sources,
            "caution": (
                "추천은 2026학년도 신입생(현 고1) 편제표와 커리어넷·대교협 공개 자료 기준입니다. "
                "실제 개설 여부와 선택 규칙(택N)은 학교 안내를 최종 확인하세요."
            ),
        }

    # -------------------------------------------------------------- layer 1
    def _layer1_dept(self, major: str, track: str) -> dict[str, Any]:
        recs = self.store.dept_recs.get(major, {})
        track_rates = self.store.rate_by_track.get(track, {})
        categories = []
        for kind in ("일반", "진로", "융합"):
            subjects = []
            seen = set()
            for subject in recs.get(kind, []):
                key = norm(subject)
                if key in seen:
                    continue
                seen.add(key)
                subjects.append({"subject": subject, "track_rate": track_rates.get(subject)})
            subjects.sort(key=lambda s: -(s["track_rate"] or 0))
            categories.append({"kind": kind, "subjects": subjects})
        return {
            "dept": major,
            "track": track,
            "categories": categories,
            "total": sum(len(c["subjects"]) for c in categories),
        }

    # -------------------------------------------------------------- layer 2
    def _expand_track_subject(self, text: str) -> tuple[list[str], list[str], list[str]]:
        """대학트랙 과목명을 (개별 공식과목들, 교과군 포괄 표현들, 폐기된 원문들)로 분해한다.

        예: '기하 또는 미적분Ⅱ' -> (['기하', '미적분Ⅱ'], [], []) / '과학' -> ([], ['과학'], [])
        마스터에도 GROUP_TERMS에도 걸리지 않는 잔여는 가짜 과목으로 남기지 않고 폐기한다
        (감사 결과 마스터-정합 과목은 전부 매칭되므로 안전. 폐기 건은 layer2.dropped_terms로 노출).
        """
        subjects, groups, dropped = [], [], []
        for part in re.split(r"\s*(?:또는|,|/)\s*", text):
            part = part.strip()
            if not part:
                continue
            for expanded in L2_SUBJECT_ALIASES.get(part, [part]):
                key = norm(expanded)
                official = self.store.subject_master.get(key)
                if official:
                    subjects.append(official)
                elif key in GROUP_TERMS:
                    groups.append(part)
                else:
                    dropped.append(expanded)
        return subjects, groups, dropped

    def _layer2_universities(self, major: str) -> dict[str, Any]:
        agg: dict[str, dict[str, Any]] = {}
        group_mentions: dict[str, set] = defaultdict(set)
        dropped_counts: dict[str, int] = defaultdict(int)
        matched_units = set()
        for row in self.store.uni_tracks:
            if row.subject.startswith("("):  # (계열 공통) 등 메타 행
                continue
            if self._major_similarity(major, row.unit) < 70:
                continue
            matched_units.add(f"{row.university} {row.unit}")
            subjects, groups, dropped = self._expand_track_subject(row.subject)
            for group in groups:
                group_mentions[group].add(row.university)
            for term in dropped:
                dropped_counts[term] += 1
            for subject in subjects:
                item = agg.setdefault(
                    subject,
                    {"subject": subject, "universities": set(), "types": set(), "comment": ""},
                )
                item["universities"].add(row.university)
                item["types"].add(row.track_type)
                if row.comment and not item["comment"]:
                    item["comment"] = row.comment

        dept_rates = self.store.rate_by_dept.get(major, {})
        subjects = []
        for item in agg.values():
            unis = sorted(item["universities"])
            subjects.append(
                {
                    "subject": item["subject"],
                    "type": "핵심" if "핵심" in item["types"] else "권장",
                    "university_count": len(unis),
                    "universities": unis[:6],
                    "rate": dept_rates.get(item["subject"]),
                    "comment": item["comment"][:120],
                }
            )
        subjects.sort(key=lambda s: (0 if s["type"] == "핵심" else 1, -s["university_count"]))
        return {
            "unit_count": len(matched_units),
            "units_sample": sorted(matched_units)[:8],
            "subjects": subjects[:20],
            "group_mentions": sorted(
                ({"group": g, "university_count": len(unis)} for g, unis in group_mentions.items()),
                key=lambda x: -x["university_count"],
            ),
            "dropped_terms": sorted(
                ({"term": term, "count": count} for term, count in dropped_counts.items()),
                key=lambda x: -x["count"],
            ),
        }

    # -------------------------------------------------------------- layer 3
    def _layer3_school(self, school: str, layer1: dict, layer2: dict) -> dict[str, Any]:
        offerings = self.store.schools.get(school, [])
        offer_map: dict[str, list[Offering]] = defaultdict(list)
        for item in offerings:
            offer_map[norm(item.subject)].append(item)

        # 추천 과목 통합: L2 핵심 > L2 권장 > L1 진로 > L1 일반 > L1 융합
        merged: dict[str, dict[str, Any]] = {}

        def add(subject: str, source: str, rank: int, extra: dict):
            key = norm(subject)
            if key in GROUP_TERMS:  # 교과군 포괄 표현은 개별 과목 대조 대상이 아님
                return
            if key not in merged or rank < merged[key]["rank"]:
                merged[key] = {"subject": subject, "source": source, "rank": rank, **extra}

        for item in layer2["subjects"]:
            rank = 0 if item["type"] == "핵심" else 1
            add(item["subject"], f"대학 {item['type']}", rank,
                {"university_count": item["university_count"], "rate": item["rate"]})
        kind_rank = {"진로": 2, "일반": 3, "융합": 4}
        for category in layer1["categories"]:
            for entry in category["subjects"]:
                add(entry["subject"], f"커리어넷 {category['kind']}", kind_rank[category["kind"]],
                    {"track_rate": entry["track_rate"]})

        available, unavailable = [], []
        for info in sorted(merged.values(), key=lambda x: x["rank"]):
            rows = offer_map.get(norm(info["subject"]))
            if rows:
                semesters = sorted({s for r in rows for s in r.semesters})
                # 같은 과목이 여러 행이면 지정(택N 그룹 미소속) 우선(감사상 충돌 0건이라 실질 단일)
                plain_rows = [r for r in rows if not r.choice_group]
                if plain_rows:
                    mode, take_n, choice_group = "지정", 0, ""
                else:
                    rep = rows[0]
                    mode, take_n, choice_group = "선택군", rep.take_n, rep.choice_group
                available.append(
                    {
                        **{k: v for k, v in info.items() if k != "rank"},
                        "semesters": semesters,
                        "semester_labels": [SEMESTER_LABELS[s] for s in semesters],
                        "sections": sorted({r.section for r in rows if r.section}),
                        "mode": mode,
                        "take_n": take_n,
                        "choice_group": choice_group,
                    }
                )
            else:
                scope = "공통 미개설" if norm(info["subject"]) not in self.store.all_offered_keys else ""
                unavailable.append({**{k: v for k, v in info.items() if k != "rank"}, "scope": scope})

        # 학기별 정리(로드맵 뷰): 2학년 이후 학기 중심
        by_semester: dict[str, list[str]] = defaultdict(list)
        for item in available:
            for sem in item["semesters"]:
                by_semester[sem].append(item["subject"])
        plan = [
            {"semester": sem, "label": SEMESTER_LABELS[sem], "subjects": by_semester[sem][:8]}
            for sem in SEMESTER_KEYS if by_semester.get(sem)
        ]

        return {
            "school": school,
            "offering_count": len(offerings),
            "available": available[:24],
            "unavailable": unavailable[:12],
            "plan": plan,
        }

    # ---------------------------------------------------------- matching
    def _pick_school(self, text: Any) -> str:
        text_norm = norm(text)
        if not text_norm:
            return ""
        for school in self.store.schools:
            if norm(school) in text_norm or text_norm in norm(school):
                return school
        return ""

    def _pick_major(self, text: Any) -> str:
        repaired = repair_text(text)
        matched = self._match_major_text(repaired)
        if matched:
            return matched
        replacements = {
            "컴공": "컴퓨터공학", "컴퓨터": "컴퓨터공학", "소프트웨어": "컴퓨터공학",
            "기계": "기계공학", "생명": "생명공학", "바이오": "생명공학",
        }
        for key, value in replacements.items():
            if key in repaired:
                matched = self._match_major_text(value)
                if matched:
                    return matched
                break
        return self._best_major_by_overlap(repaired)

    def _match_major_text(self, text: str) -> str:
        text_norm = norm(text)
        if not text_norm:
            return ""
        base = self._major_base_key(text)
        for dept in self.store.dept_recs:
            if norm(dept) == text_norm or self._major_base_key(dept) == base:
                return dept
        candidates = []
        for dept in self.store.dept_recs:
            dept_norm = norm(dept)
            if dept_norm in text_norm or text_norm in dept_norm:
                candidates.append(dept)
        if candidates:
            # '경영학' -> '의료경영학과'가 아니라 '경영학과'를 고르도록 가장 짧은 후보 선택
            return min(candidates, key=lambda d: len(norm(d)))
        return ""

    def _best_major_by_overlap(self, text: str) -> str:
        tokens = set(re.findall(r"[가-힣A-Za-z0-9]+", repair_text(text)))
        best = ("", 0)
        for dept in self.store.dept_recs:
            score = sum(1 for token in tokens if token and token in dept)
            if score > best[1]:
                best = (dept, score)
        return best[0] if best[1] else ""

    def _major_base_key(self, text: str) -> str:
        value = norm(text)
        for suffix in ("학과", "전공", "계열", "학부", "과", "부"):
            value = re.sub(f"{suffix}$", "", value)
        return value

    def _major_similarity(self, target: str, candidate: str) -> int:
        t = self._major_base_key(target)
        c = self._major_base_key(candidate)
        if not t or not c:
            return 0
        if t == c:
            return 100
        if len(t) >= 2 and len(c) >= 2 and (t in c or c in t):
            return 80
        aliases = {
            "컴퓨터공학": ["컴퓨터", "소프트웨어", "ai", "인공지능", "정보컴퓨터", "데이터"],
            "생명공학": ["생명", "바이오", "의생명", "식품생명", "분자생명"],
            "기계공학": ["기계", "로봇", "자동차", "메카트로닉스"],
        }
        for root, words in aliases.items():
            if norm(root) in norm(target) and any(norm(word) in norm(candidate) for word in words):
                return 70
        return 0

    # ---------------------------------------------------------- output
    def _graph_payload(self, school: str, major: str, layer3: dict) -> dict[str, Any]:
        nodes = [
            {"id": f"school:{school}", "label": school, "type": "school"},
            {"id": f"major:{major}", "label": major, "type": "major"},
        ]
        edges = []
        for item in layer3["available"][:8]:
            sid = f"subject:{item['subject']}"
            nodes.append({"id": sid, "label": item["subject"], "type": "available"})
            edges.append({"source": f"school:{school}", "target": sid, "label": "개설"})
            edges.append({"source": sid, "target": f"major:{major}", "label": item["source"]})
        for item in layer3["unavailable"][:4]:
            sid = f"subject:{item['subject']}"
            nodes.append({"id": sid, "label": item["subject"], "type": "missing"})
            edges.append({"source": sid, "target": f"major:{major}", "label": item["source"]})
        return {"nodes": nodes, "edges": edges}

    def _compose_answer(self, result: dict[str, Any]) -> str:
        school, major = result["school"], result["major"]
        layer1, layer2, layer3 = result["layer1"], result["layer2"], result["layer3"]
        lines = []

        # 1. 커리어넷 포괄 추천
        parts = []
        for category in layer1["categories"]:
            names = [s["subject"] for s in category["subjects"][:5]]
            if names:
                parts.append(f"{category['kind']} 선택은 {', '.join(names)}")
        track_note = f" ({layer1['track']} 기준)" if layer1["track"] else ""
        lines.append(
            f"학과 추천 과목{track_note}\n커리어넷 기준 {major} 관련 " + " / ".join(parts)
            if parts else f"학과 추천 과목\n{major}에 대한 커리어넷 추천 데이터를 찾지 못했습니다."
        )

        # 2. 대학별 추가 추천
        if layer2["subjects"]:
            core = [s for s in layer2["subjects"] if s["type"] == "핵심"][:5]
            rec = [s for s in layer2["subjects"] if s["type"] == "권장"][:4]
            uni_lines = [f"대학별 강조 과목\n대교협 시행계획 기준 관련 모집단위 {layer2['unit_count']}곳을 보면,"]
            if core:
                core_parts = []
                for s in core:
                    rate_note = ", 제시율 {:.0f}%".format(s["rate"]) if s["rate"] else ""
                    core_parts.append(f"{s['subject']}({s['university_count']}개 대학{rate_note})")
                uni_lines.append("핵심 과목: " + ", ".join(core_parts))
            if rec:
                uni_lines.append("권장 과목: " + ", ".join(f"{s['subject']}({s['university_count']}개 대학)" for s in rec))
            groups = layer2.get("group_mentions", [])[:3]
            if groups:
                uni_lines.append(
                    "이외에 " + ", ".join(f"{g['group']} 교과 전반({g['university_count']}개 대학)" for g in groups)
                    + "을 폭넓게 강조합니다."
                )
            lines.append("\n".join(uni_lines))
        else:
            lines.append("대학별 강조 과목\n대교협 자료에서 직접 일치하는 모집단위를 찾지 못했습니다. 학과명을 조금 바꿔 검색해 보세요.")

        # 3. 학교 편제표 대조
        available = layer3["available"]
        if available:
            top = available[:6]
            avail_lines = [
                f"{school} 편제표 확인\n위 추천 과목 중 {school}에서 실제 선택할 수 있는 과목은 다음과 같습니다."
            ]
            for item in top:
                sems = ", ".join(item["semester_labels"]) or "학기 미확정"
                if item.get("mode") == "선택군":
                    mode_note = f" (택{item['take_n']} 선택군 — 선택 신청 필요)" if item.get("take_n") else " (선택군 — 선택 신청 필요)"
                else:
                    mode_note = " (학교지정)"
                avail_lines.append(f"- {item['subject']} [{item['source']}]{mode_note} → {sems}")
            lines.append("\n".join(avail_lines))
        else:
            lines.append(f"{school} 편제표 확인\n추천 과목 중 {school} 편제표에서 확인되는 과목이 없습니다. 편제표 데이터를 점검해 주세요.")

        unavailable_common = [item for item in layer3["unavailable"] if item.get("scope") == "공통 미개설"]
        unavailable_normal = [item for item in layer3["unavailable"] if item.get("scope") != "공통 미개설"]
        if unavailable_normal:
            names = ", ".join(item["subject"] for item in unavailable_normal[:5])
            lines.append(
                f"미개설 과목\n{names}은(는) 추천 근거에는 있지만 {school} 편제표에서 확인되지 않습니다. "
                "공동교육과정이나 온라인학교 개설 여부를 확인해 보세요."
            )
        if unavailable_common:
            names = ", ".join(item["subject"] for item in unavailable_common[:5])
            lines.append(
                f"일반고 공통 미개설\n{names}은(는) 조사된 모든 학교 편제표에서 확인되지 않습니다. "
                "일반고 공통 미개설 — 공동교육과정·온라인학교 확인이 필요합니다."
            )

        lines.append(f"주의\n{result['caution']}")
        return "\n\n".join(lines)

    def _need_more(self, message: str) -> dict[str, Any]:
        return {
            "mode": "need_more",
            "answer": message,
            "layer1": {"dept": "", "track": "", "categories": [], "total": 0},
            "layer2": {"unit_count": 0, "units_sample": [], "subjects": [], "group_mentions": [], "dropped_terms": []},
            "layer3": {"school": "", "offering_count": 0, "available": [], "unavailable": [], "plan": []},
            "graph": {"nodes": [], "edges": []},
            "sources": [],
            "caution": "",
        }
