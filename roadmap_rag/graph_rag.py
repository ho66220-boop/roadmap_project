from __future__ import annotations

import csv
from dataclasses import dataclass, field
from pathlib import Path
import re
from typing import Any

import openpyxl


SEMESTER_COLUMNS = [
    ("1학년 1학기", 7),
    ("1학년 2학기", 8),
    ("2학년 1학기", 9),
    ("2학년 2학기", 10),
    ("3학년 1학기", 11),
    ("3학년 2학기", 12),
]


@dataclass(frozen=True)
class SubjectOffering:
    name: str
    group: str
    subject_type: str
    semesters: tuple[str, ...]
    credits: float = 0.0


@dataclass(frozen=True)
class MajorSubject:
    major: str
    subject_code: str
    subject: str
    group: str
    subject_type: str
    recommend_type: str
    priority: float
    university_count: float
    evidence: str


@dataclass(frozen=True)
class AdmissionResult:
    result_year: str
    university: str
    major: str
    area: str
    admission_period: str
    admission_type: str
    track_name: str
    recruit_count: str
    additional_pass_count: str
    competition_rate: str
    grade_50: float | None
    grade_70: float | None
    score_50: str
    score_70: str
    source_url: str


@dataclass
class RoadmapGraph:
    schools: dict[str, list[SubjectOffering]] = field(default_factory=dict)
    majors: dict[str, list[MajorSubject]] = field(default_factory=dict)
    subject_aliases: dict[str, str] = field(default_factory=dict)
    admissions: list[AdmissionResult] = field(default_factory=list)
    careernet_subject_count: int = 0


def repair_text(value: Any) -> str:
    """Repair UTF-8 text that was accidentally decoded as latin-1/cp1252."""
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    if any(marker in text for marker in ("ì", "í", "ê", "ë", "â")):
        for encoding in ("latin1", "cp1252"):
            try:
                repaired = text.encode(encoding).decode("utf-8")
            except UnicodeError:
                continue
            if repaired and repaired != text:
                return repaired.strip()
    return text


def norm(value: Any) -> str:
    text = repair_text(value).lower()
    text = text.replace("Ⅰ", "1").replace("Ⅱ", "2")
    text = re.sub(r"[\s·ㆍ․,/_()\-\[\]]+", "", text)
    text = text.replace("대학교", "대")
    return text


def to_float(value: Any, default: float = 0.0) -> float:
    try:
        if value in (None, "", "-"):
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def optional_float(value: Any) -> float | None:
    try:
        text = str(value).strip()
        if text in {"", "-", "0"}:
            return None
        return float(text)
    except (TypeError, ValueError):
        return None


def split_subjects(text: str) -> list[str]:
    raw = re.split(r"[,/|]+|\n|ㆍ|·", text or "")
    return [repair_text(item).strip() for item in raw if repair_text(item).strip()]


class GraphRAGEngine:
    def __init__(self, workbook_path: Path, admission_path: Path | None = None, careernet_path: Path | None = None):
        project_root = Path(__file__).resolve().parents[1]
        self.workbook_path = workbook_path
        self.admission_path = admission_path or project_root / "data" / "admission_results_core.csv"
        self.careernet_path = careernet_path or project_root / "data" / "careernet_major_subjects.csv"
        self.graph = self._load_graph(workbook_path)

    @classmethod
    def from_default_workbook(cls) -> "GraphRAGEngine":
        root = Path(__file__).resolve().parents[1]
        sample = root / "data" / "sample_roadmap.xlsx"
        return cls(sample if sample.exists() else root.parent / "울산고교_과목로드맵_수정.xlsx")

    def meta(self) -> dict[str, Any]:
        return {
            "workbook": self.workbook_path.name,
            "admission_file": self.admission_path.name if self.admission_path.exists() else "",
            "careernet_file": self.careernet_path.name if self.careernet_path.exists() else "",
            "school_count": len(self.graph.schools),
            "major_count": len(self.graph.majors),
            "admission_result_count": len(self.graph.admissions),
            "careernet_subject_count": self.graph.careernet_subject_count,
            "schools": sorted(self.graph.schools)[:200],
            "majors": sorted(self.graph.majors)[:300],
        }

    def chat(self, message: str, profile: dict[str, Any] | None = None) -> dict[str, Any]:
        profile = profile or {}
        school = self._pick_school(profile.get("school") or message)
        major = self._pick_major(profile.get("major") or message)
        grade = self._pick_grade(profile.get("grade") or message)
        taken = self._pick_taken(profile.get("taken") or message)
        school_grade = self._pick_school_grade(profile.get("school_grade") or profile.get("gpa") or message)
        admission_focus = self._pick_admission_focus(profile.get("admission_focus") or message)

        if not school:
            return self._need_more("학교명을 알려주면 해당 학교 편제표 안에서 실제 선택 가능한 과목만 골라볼 수 있어요.")
        if not major:
            return self._need_more("목표 학과나 계열을 알려주면 대학트랙 기준 핵심/권장 과목을 찾아볼게요.")

        recommendation = self.recommend(
            school=school,
            major=major,
            grade=grade,
            taken_subjects=taken,
            school_grade=school_grade,
            admission_focus=admission_focus,
        )
        recommendation["answer"] = self._compose_answer(recommendation)
        return recommendation

    def recommend(
        self,
        school: str,
        major: str,
        grade: int = 1,
        taken_subjects: list[str] | None = None,
        school_grade: float | None = None,
        admission_focus: str = "전체",
    ) -> dict[str, Any]:
        taken_subjects = taken_subjects or []
        offerings = self.graph.schools.get(school, [])
        major_subjects = self.graph.majors.get(major, [])
        school_by_norm = {norm(item.name): item for item in offerings}
        taken_norms = {norm(item) for item in taken_subjects}

        ranked = self._rank_major_subjects(major_subjects)
        available = []
        unavailable = []
        completed = []

        for item in ranked:
            key = norm(item.subject)
            offering = school_by_norm.get(key)
            row = {
                "subject": item.subject,
                "recommend_type": item.recommend_type,
                "priority": item.priority,
                "university_count": item.university_count,
                "evidence": item.evidence,
                "group": item.group,
                "semesters": list(offering.semesters) if offering else [],
            }
            if key in taken_norms:
                completed.append(row)
            elif offering:
                available.append(row)
            else:
                unavailable.append(row)

        admission = self._recommend_admissions(major=major, school_grade=school_grade, admission_focus=admission_focus)
        plan = self._build_plan(available, grade)
        similar_majors = self._similar_majors(major)

        return {
            "mode": "recommendation",
            "school": school,
            "major": major,
            "grade": grade,
            "school_grade": school_grade,
            "admission_focus": admission_focus,
            "taken_subjects": taken_subjects,
            "completed": completed[:8],
            "available": available[:10],
            "unavailable": unavailable[:8],
            "plan": plan,
            "similar_majors": similar_majors,
            "admission": admission,
            "graph": self._graph_payload(school, major, available[:6], unavailable[:4], admission),
            "sources": [
                f"{self.workbook_path.name} / 학교편제표",
                f"{self.workbook_path.name} / 로드맵DB_v2_대학트랙집계",
                *(
                    ["data/careernet_major_subjects.csv / 커리어넷 2022 개정 선택과목"]
                    if self.graph.careernet_subject_count
                    else []
                ),
                *(
                    ["data/admission_results_core.csv / 대학어디가 입시결과 샘플"]
                    if self.graph.admissions
                    else []
                ),
            ],
            "caution": (
                "입시결과는 대학어디가 공개 자료 샘플 기준이며 대학별 환산 방식이 다릅니다. "
                "50%/70% cut은 합격 보장이 아니라 참고 지표입니다."
            ),
        }

    def _load_graph(self, workbook_path: Path) -> RoadmapGraph:
        if not workbook_path.exists():
            raise FileNotFoundError(
                f"원본 로드맵 엑셀 파일을 찾을 수 없습니다. README의 데이터 준비 방법을 확인하세요: {workbook_path}"
            )

        workbook = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
        graph = RoadmapGraph()
        self._load_aliases(workbook, graph)
        self._load_school_offerings(workbook, graph)
        self._load_major_subjects(workbook, graph)
        self._load_careernet_subjects(graph)
        self._load_admissions(graph)
        return graph

    def _load_aliases(self, workbook: Any, graph: RoadmapGraph) -> None:
        if "과목별칭" not in workbook.sheetnames:
            return
        sheet = workbook["과목별칭"]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            raw = repair_text(row[1] if len(row) > 1 else "")
            normalized = repair_text(row[2] if len(row) > 2 else "")
            if raw and normalized:
                graph.subject_aliases[norm(raw)] = normalized

    def _load_school_offerings(self, workbook: Any, graph: RoadmapGraph) -> None:
        sheet = workbook["학교편제표"]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) < 5:
                continue
            school = repair_text(row[0])
            subject = repair_text(row[4])
            if not school or not subject:
                continue
            semesters = []
            for label, idx in SEMESTER_COLUMNS:
                if idx < len(row) and to_float(row[idx]) > 0:
                    semesters.append(label)
            offering = SubjectOffering(
                name=subject,
                group=repair_text(row[2]),
                subject_type=repair_text(row[3]),
                semesters=tuple(semesters),
                credits=to_float(row[6] if len(row) > 6 else 0),
            )
            graph.schools.setdefault(school, []).append(offering)

    def _load_major_subjects(self, workbook: Any, graph: RoadmapGraph) -> None:
        sheet = workbook["로드맵DB_v2_대학트랙집계"]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) < 12:
                continue
            major = repair_text(row[1])
            subject_code = repair_text(row[2])
            subject = repair_text(row[3])
            if not major or not subject:
                continue
            if subject_code in {"UNMAPPED"} or subject_code.startswith("BROAD-"):
                continue
            item = MajorSubject(
                major=major,
                subject_code=subject_code,
                subject=subject,
                group=repair_text(row[4]),
                subject_type=repair_text(row[5]),
                recommend_type=repair_text(row[6]) or "참고",
                priority=to_float(row[7], 99.0),
                university_count=to_float(row[8], 0.0),
                evidence=repair_text(row[11]),
            )
            graph.majors.setdefault(major, []).append(item)

    def _load_careernet_subjects(self, graph: RoadmapGraph) -> None:
        if not self.careernet_path.exists():
            return
        category_fields = [
            ("general_selection_subjects", "권장", "일반 선택", 40.0),
            ("career_selection_subjects", "핵심", "진로 선택", 20.0),
            ("convergence_selection_subjects", "권장", "융합 선택", 60.0),
        ]
        with self.careernet_path.open(encoding="utf-8-sig", newline="") as file:
            for row in csv.DictReader(file):
                career_major = repair_text(row.get("career_major_name"))
                if not career_major:
                    continue
                major = self._match_graph_major(career_major, graph)
                source_url = repair_text(row.get("source_url"))
                for field, recommend_type, category, priority in category_fields:
                    for subject in self._split_careernet_subjects(row.get(field, "")):
                        graph.majors.setdefault(major, []).append(
                            MajorSubject(
                                major=major,
                                subject_code=f"CAREERNET-{row.get('career_seq', '')}-{field}",
                                subject=subject,
                                group="커리어넷",
                                subject_type=category,
                                recommend_type=recommend_type,
                                priority=priority,
                                university_count=0.0,
                                evidence=(
                                    f"커리어넷 {career_major}의 2022 개정 교육과정 {category} 관련 과목"
                                    + (f" ({source_url})" if source_url else "")
                                ),
                            )
                        )
                        graph.careernet_subject_count += 1

    def _split_careernet_subjects(self, text: Any) -> list[str]:
        subjects = []
        seen = set()
        for item in re.split(r"\|", repair_text(text)):
            subject = repair_text(item).strip()
            if not subject:
                continue
            if len(subject) == 1:
                continue
            key = norm(subject)
            if key in seen:
                continue
            seen.add(key)
            subjects.append(subject)
        return subjects

    def _match_graph_major(self, career_major: str, graph: RoadmapGraph) -> str:
        career_key = self._major_base_key(career_major)
        for major in graph.majors:
            if self._major_base_key(major) == career_key:
                return major
        for major in graph.majors:
            major_key = self._major_base_key(major)
            if len(major_key) >= 2 and len(career_key) >= 2 and (major_key in career_key or career_key in major_key):
                return major
        return career_major

    def _major_base_key(self, text: str) -> str:
        value = norm(text)
        for suffix in ("학과", "전공", "계열", "학부", "과", "부"):
            value = re.sub(f"{suffix}$", "", value)
        return value

    def _load_admissions(self, graph: RoadmapGraph) -> None:
        if not self.admission_path.exists():
            return
        with self.admission_path.open(encoding="utf-8-sig", newline="") as file:
            for row in csv.DictReader(file):
                grade_70 = optional_float(row.get("grade_70"))
                grade_50 = optional_float(row.get("grade_50"))
                if grade_70 is None and grade_50 is None:
                    continue
                graph.admissions.append(
                    AdmissionResult(
                        result_year=row.get("result_year", ""),
                        university=row.get("university", ""),
                        major=row.get("major", ""),
                        area=row.get("area", ""),
                        admission_period=row.get("admission_period", ""),
                        admission_type=row.get("admission_type", ""),
                        track_name=row.get("track_name", ""),
                        recruit_count=row.get("recruit_count", ""),
                        additional_pass_count=row.get("additional_pass_count", ""),
                        competition_rate=row.get("competition_rate", ""),
                        grade_50=grade_50,
                        grade_70=grade_70,
                        score_50=row.get("score_50", ""),
                        score_70=row.get("score_70", ""),
                        source_url=row.get("source_url", ""),
                    )
                )

    def _recommend_admissions(self, major: str, school_grade: float | None, admission_focus: str = "전체") -> dict[str, Any]:
        if not self.graph.admissions:
            return {"available": False, "message": "입시결과 CSV가 없어 대학 추천은 비활성화되어 있습니다.", "exact": [], "similar": [], "reach": [], "target": [], "likely": [], "by_type": {}}
        candidates = []
        for result in self.graph.admissions:
            similarity = self._major_similarity(major, result.major)
            if similarity <= 0:
                continue
            cutoff = result.grade_70 or result.grade_50
            if cutoff is None:
                continue
            band = self._admission_band(school_grade, cutoff)
            direct = self._is_direct_major_name(major, result.major)
            candidates.append(
                {
                    "university": result.university,
                    "major": result.major,
                    "area": result.area,
                    "admission_period": result.admission_period,
                    "admission_type": result.admission_type,
                    "track_name": result.track_name,
                    "competition_rate": result.competition_rate,
                    "grade_50": result.grade_50,
                    "grade_70": result.grade_70,
                    "cutoff": cutoff,
                    "source_url": result.source_url,
                    "similarity": similarity,
                    "match_type": "동일/직접 관련" if direct else "대체 유사 학과",
                    "direct": direct,
                    "band": band,
                    "category": self._admission_category(result),
                }
            )

        def sort_key(item: dict[str, Any]) -> tuple[int, float, float]:
            band_order = {"안정": 0, "적정": 1, "상향": 2, "도전": 3, "참고": 4}
            return (band_order.get(item["band"], 9), -item["similarity"], item["cutoff"])

        ranked = sorted(candidates, key=sort_key)
        by_type = {
            category: self._admission_bucket([item for item in ranked if item["category"] == category])
            for category in ("교과", "종합", "기타")
        }
        selected = ranked if admission_focus == "전체" else [item for item in ranked if item["category"] == admission_focus]
        exact = [item for item in selected if item["direct"]]
        similar = [item for item in selected if not item["direct"]]
        return {
            "available": True,
            "student_grade": school_grade,
            "focus": admission_focus,
            "by_type": by_type,
            "exact": exact[:8],
            "similar": similar[:8],
            "likely": [item for item in selected if item["band"] == "안정"][:5],
            "target": [item for item in selected if item["band"] == "적정"][:5],
            "reach": [item for item in selected if item["band"] == "상향"][:5],
            "count": len(candidates),
        }

    def _admission_bucket(self, items: list[dict[str, Any]]) -> dict[str, Any]:
        return {
            "exact": [item for item in items if item["direct"]][:8],
            "similar": [item for item in items if not item["direct"]][:8],
            "likely": [item for item in items if item["band"] == "안정"][:5],
            "target": [item for item in items if item["band"] == "적정"][:5],
            "reach": [item for item in items if item["band"] == "상향"][:5],
            "count": len(items),
        }

    def _admission_category(self, result: AdmissionResult) -> str:
        text = f"{result.admission_type} {result.track_name}"
        if "교과" in text:
            return "교과"
        if "종합" in text:
            return "종합"
        return "기타"

    def _admission_band(self, student_grade: float | None, cutoff: float) -> str:
        if student_grade is None:
            return "참고"
        # Lower Korean school grade is stronger. Compare against 70% cut.
        diff = student_grade - cutoff
        if diff <= -0.4:
            return "안정"
        if diff <= 0.25:
            return "적정"
        if diff <= 0.8:
            return "상향"
        return "도전"

    def _major_similarity(self, target: str, candidate: str) -> int:
        t = norm(target)
        c = norm(candidate)
        if not t or not c:
            return 0
        if t in c or c in t:
            return 100
        aliases = {
            "컴퓨터공학": ["컴퓨터", "소프트웨어", "ai", "인공지능", "정보컴퓨터"],
            "생명공학": ["생명", "바이오", "의생명", "식품생명", "분자생명"],
            "기계공학": ["기계", "로봇", "자동차", "메카트로닉스"],
        }
        for root, words in aliases.items():
            if norm(root) in t and any(norm(word) in c for word in words):
                return 70
        t_tokens = set(re.findall(r"[가-힣A-Za-z0-9]+", target))
        c_tokens = set(re.findall(r"[가-힣A-Za-z0-9]+", candidate))
        overlap = t_tokens & c_tokens
        return min(60, 20 * len(overlap)) if overlap else 0

    def _is_direct_major_name(self, target: str, candidate: str) -> bool:
        def base(text: str) -> str:
            value = norm(text)
            for suffix in ("전공", "계열", "과", "부"):
                value = re.sub(f"{suffix}$", "", value)
            return value

        target_base = base(target)
        candidate_base = base(candidate)
        if not target_base or not candidate_base:
            return False
        return candidate_base == target_base

    def _pick_school(self, text: Any) -> str:
        text_norm = norm(text)
        if not text_norm:
            return ""
        for school in self.graph.schools:
            if norm(school) in text_norm or text_norm in norm(school):
                return school
        return ""

    def _pick_major(self, text: Any) -> str:
        repaired = repair_text(text)
        matched = self._match_major_text(repaired)
        if matched:
            return matched

        replacements = {
            "컴공": "컴퓨터공학",
            "컴퓨터": "컴퓨터공학",
            "소프트웨어": "컴퓨터공학",
            "기계": "기계공학",
            "생명": "생명공학",
            "바이오": "생명공학",
        }
        for key, value in replacements.items():
            if key in repaired:
                matched = self._match_major_text(value)
                if matched:
                    return matched
                break

        return self._best_major_by_overlap(repaired)

    def _match_major_text(self, text: str) -> str:
        text_norm = norm(text)
        if not text_norm:
            return ""
        for major in self.graph.majors:
            major_norm = norm(major)
            if major_norm == text_norm:
                return major
        for major in self.graph.majors:
            major_norm = norm(major)
            if major_norm in text_norm or text_norm in major_norm:
                return major
        return ""

    def _best_major_by_overlap(self, text: str) -> str:
        tokens = set(re.findall(r"[가-힣A-Za-z0-9]+", repair_text(text)))
        best = ("", 0)
        for major in self.graph.majors:
            score = sum(1 for token in tokens if token and token in major)
            if score > best[1]:
                best = (major, score)
        return best[0] if best[1] else ""

    def _pick_grade(self, text: Any) -> int:
        if isinstance(text, int):
            return max(1, min(3, text))
        match = re.search(r"([123])\s*학년|고\s*([123])|예비\s*고\s*([123])", repair_text(text))
        if not match:
            return 1
        return int(next(group for group in match.groups() if group))

    def _pick_school_grade(self, text: Any) -> float | None:
        if isinstance(text, (int, float)):
            value = float(text)
            return value if 1 <= value <= 9 else None
        repaired = repair_text(text)
        match = re.search(r"(?:내신|등급|평균)\s*([1-9](?:\.\d+)?)", repaired)
        if not match:
            match = re.search(r"\b([1-9](?:\.\d+)?)\s*등급", repaired)
        if not match:
            return None
        value = float(match.group(1))
        return value if 1 <= value <= 9 else None

    def _pick_admission_focus(self, text: Any) -> str:
        repaired = repair_text(text)
        if repaired in {"교과", "종합", "기타"}:
            return repaired
        if "교과" in repaired:
            return "교과"
        if "종합" in repaired or "학종" in repaired:
            return "종합"
        return "전체"

    def _pick_taken(self, text: Any) -> list[str]:
        if isinstance(text, list):
            return [repair_text(item) for item in text if repair_text(item)]
        return split_subjects(repair_text(text))

    def _rank_major_subjects(self, subjects: list[MajorSubject]) -> list[MajorSubject]:
        weight = {"핵심": 0, "권장": 1, "참고": 2}
        seen = set()
        ranked = []
        for item in sorted(subjects, key=lambda x: (weight.get(x.recommend_type, 3), -x.university_count, x.priority)):
            key = norm(item.subject)
            if key in seen:
                continue
            seen.add(key)
            ranked.append(item)
        return ranked

    def _build_plan(self, available: list[dict[str, Any]], grade: int) -> list[dict[str, Any]]:
        current_order = max(1, min(6, (grade - 1) * 2 + 1))
        plan: dict[str, list[dict[str, Any]]] = {}
        for item in available:
            future_semesters = [
                semester for semester in item["semesters"]
                if self._semester_order(semester) >= current_order
            ]
            semester = future_semesters[0] if future_semesters else (item["semesters"][0] if item["semesters"] else "편제표 확인 필요")
            plan.setdefault(semester, []).append(item)

        ordered = []
        for semester, items in sorted(plan.items(), key=lambda pair: self._semester_order(pair[0])):
            ordered.append(
                {
                    "semester": semester,
                    "subjects": [item["subject"] for item in items[:4]],
                    "reason": f"{items[0]['recommend_type']} 과목을 우선 배치했습니다. 근거: {items[0]['evidence']}",
                }
            )
        return ordered[:5]

    def _semester_order(self, semester: str) -> int:
        match = re.search(r"([123])학년\s*([12])학기", semester)
        if not match:
            return 99
        return (int(match.group(1)) - 1) * 2 + int(match.group(2))

    def _similar_majors(self, major: str) -> list[str]:
        candidates = []
        for candidate in sorted(self.graph.majors):
            if candidate == major:
                continue
            score = self._major_similarity(major, candidate)
            if score > 0:
                candidates.append((score, candidate))
        return [candidate for _, candidate in sorted(candidates, reverse=True)[:5]]

    def _graph_payload(
        self,
        school: str,
        major: str,
        available: list[dict[str, Any]],
        unavailable: list[dict[str, Any]],
        admission: dict[str, Any],
    ) -> dict[str, Any]:
        nodes = [
            {"id": f"school:{school}", "label": school, "type": "school"},
            {"id": f"major:{major}", "label": major, "type": "major"},
        ]
        edges = []
        for item in available:
            sid = f"subject:{item['subject']}"
            nodes.append({"id": sid, "label": item["subject"], "type": "available"})
            edges.append({"source": f"school:{school}", "target": sid, "label": "개설"})
            edges.append({"source": sid, "target": f"major:{major}", "label": item["recommend_type"]})
        for item in unavailable:
            sid = f"subject:{item['subject']}"
            nodes.append({"id": sid, "label": item["subject"], "type": "missing"})
            edges.append({"source": sid, "target": f"major:{major}", "label": item["recommend_type"]})
        for item in (admission.get("target") or admission.get("exact") or [])[:3]:
            uid = f"admission:{item['university']}:{item['major']}"
            nodes.append({"id": uid, "label": item["university"], "type": "admission"})
            edges.append({"source": f"major:{major}", "target": uid, "label": item["band"]})
        return {"nodes": nodes, "edges": edges}

    def _compose_answer(self, result: dict[str, Any]) -> str:
        available = result["available"]
        unavailable = result["unavailable"]
        completed = result["completed"]
        plan = result["plan"]
        admission = result["admission"]

        lines = [
            f"과목 로드맵\n{result['school']} {result['grade']}학년 기준으로 {result['major']} 진로를 보면, 학교 편제표 안에서 바로 연결되는 추천 과목은 {self._join_subjects([x['subject'] for x in available[:5]])}입니다.",
        ]
        if completed:
            lines.append(f"이미 입력한 과목 중에서는 {self._join_subjects([x['subject'] for x in completed[:4]])}이 목표 학과 추천 과목과 겹칩니다.")
        if unavailable:
            lines.append(f"다만 {self._join_subjects([x['subject'] for x in unavailable[:4]])}은 목표 학과 근거에는 나오지만, 현재 학교 편제표에서는 바로 확인되지 않아 대체 과목이나 실제 개설 여부 확인이 필요합니다.")
        if plan:
            first = plan[0]
            lines.append(f"우선순위는 {first['semester']}에 {self._join_subjects(first['subjects'])}부터 잡는 방식이 좋습니다.")
        if admission.get("available"):
            if result.get("school_grade"):
                lines.append(self._admission_answer(admission))
            else:
                lines.append("내신컷 해석\n내신 등급을 함께 입력하면 대학어디가 샘플 입시결과 기준으로 교과/종합 후보를 나눠볼 수 있습니다.")
        lines.append(f"주의\n{result['caution']}")
        return "\n\n".join(lines)

    def _admission_answer(self, admission: dict[str, Any]) -> str:
        focus = admission.get("focus", "전체")
        intro = "내신컷 해석"
        if focus == "전체":
            intro += "\n교과와 종합은 평가 방식이 달라서 따로 봐야 합니다. 우선 교과/종합 중 무엇을 중심으로 볼지 정하면 더 정확히 좁힐 수 있어요."
        else:
            intro += f"\n현재는 {focus} 중심으로 후보를 정리했습니다."

        sections = [intro]
        sections.append(self._admission_type_summary("교과", admission.get("by_type", {}).get("교과", {})))
        sections.append(self._admission_type_summary("종합", admission.get("by_type", {}).get("종합", {})))

        similar = admission.get("similar", [])
        if similar:
            sections.append(f"대체 유사 학과\n{self._join_university_major(similar[:4])}도 비교할 만합니다.")
        return "\n\n".join(section for section in sections if section.strip())

    def _admission_type_summary(self, category: str, bucket: dict[str, Any]) -> str:
        if not bucket or not bucket.get("count"):
            return f"{category} 기준\n현재 수집된 샘플 안에서는 직접 비교할 후보가 부족합니다."

        if category == "교과":
            guide = "교과는 내신 환산등급과 대학별 반영 교과/환산식을 먼저 봅니다. 아래 후보는 70% cut을 기준으로 단순 비교한 범위입니다."
        else:
            guide = "종합은 내신컷만으로 판단하기 어렵고, 편제표 안에서 목표 학과의 핵심 과목을 이어 듣는지와 세특·탐구 흐름을 함께 봐야 합니다."

        lines = [f"{category} 기준\n{guide}"]
        for label, key in [("안정", "likely"), ("적정", "target"), ("상향", "reach")]:
            items = bucket.get(key, [])
            if items:
                lines.append(f"{label}: {self._join_university_major(items[:3])}")
        return "\n".join(lines)

    def _join_university_major(self, items: list[dict[str, Any]]) -> str:
        return ", ".join(f"{item['university']} {item['major']}({item['grade_70'] or item['grade_50']})" for item in items)

    def _join_subjects(self, subjects: list[str]) -> str:
        return ", ".join(subjects) if subjects else "확인된 과목 없음"

    def _need_more(self, message: str) -> dict[str, Any]:
        return {
            "mode": "need_more",
            "answer": message,
            "available": [],
            "unavailable": [],
            "completed": [],
            "plan": [],
            "similar_majors": [],
            "admission": {"available": False, "exact": [], "similar": [], "reach": [], "target": [], "likely": []},
            "graph": {"nodes": [], "edges": []},
            "sources": [],
            "caution": "",
        }
